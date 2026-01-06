
# Create your views here.

import os
from boran_app.models import MovimientoUnificadoCredito, MovimientoUnificadoDebito, VentasConsulta
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from django.http import JsonResponse
from boran_app.utils import regenerar_resumenes_credito_debito
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_exempt
from boran_app.models import ResumenCredito, ResumenDebito
from .utils import poblar_movimientos_unificados_credito, poblar_movimientos_unificados_debito
from .eerr import generar_estado_resultados
from .utils_balance import obtener_matriz_balance
from .balance_utils import obtener_matriz_dict_balance
from boran_app.models import VentasConsulta
from datetime import date, datetime
from decimal import Decimal
from .models import ResumenMensual
import os
from django.shortcuts import redirect
from boran_app.models import ResumenCredito, ResumenDebito
from boran_app.utils_balance import calcular_resultados_mensuales
from .models import MovimientoUnificadoDebito, MovimientoUnificadoCredito, ResumenMensual
from django.db.models import Sum
from collections import defaultdict
from django.db.models.functions import TruncMonth
from django.shortcuts import render
from .models import ResultadoMensualDetalle

# ——————————————————————————————————————————————————————————————
#   MANEJO DE AÑO FISCAL (IGUAL QUE PANEL-BRONZ)
# ——————————————————————————————————————————————————————————————

PANEL_YEAR_CHOICES = (2025, 2026)


def _get_default_panel_year() -> int:
    """Devuelve el año por defecto (año actual si está en las opciones)."""
    today = date.today()
    if today.year in PANEL_YEAR_CHOICES:
        return today.year
    return PANEL_YEAR_CHOICES[0]


def get_panel_year(request) -> int:
    """Obtiene el año del panel desde la sesión."""
    year = request.session.get('panel_year')
    if year is None or int(year) not in PANEL_YEAR_CHOICES:
        year = _get_default_panel_year()
        request.session['panel_year'] = year
    return int(year)


def get_panel_date_range(request):
    """
    Obtiene el rango de fechas para el año seleccionado.
    Si es el año actual, end_date es hoy. Si no, es 31/12 del año.
    """
    year = get_panel_year(request)
    start_date = date(year, 1, 1)
    today = date.today()
    if year == today.year:
        end_date = today
    else:
        end_date = date(year, 12, 31)
    return year, start_date, end_date


def regenerate_financial_tables(start_date, end_date):
    """Regenera todas las tablas financieras para el rango de fechas dado."""
    from .utils import regenerar_ventas_consulta
    regenerar_ventas_consulta(start_date=start_date, end_date=end_date)
    poblar_movimientos_unificados_debito(start_date=start_date, end_date=end_date)
    poblar_movimientos_unificados_credito(start_date=start_date, end_date=end_date)
    regenerar_resumenes_credito_debito()


# Alias para compatibilidad con código existente
def obtener_fechas_anno_fiscal(request):
    """
    Alias de get_panel_date_range para compatibilidad.
    Devuelve (fecha_inicio, fecha_fin, anno_fiscal) en ese orden.
    Los valores de fecha se devuelven como objetos date.
    """
    year, start_date, end_date = get_panel_date_range(request)
    return start_date, end_date, year


def obtener_fechas_anno_fiscal_str(request):
    """
    Igual que obtener_fechas_anno_fiscal pero devuelve fechas como strings ISO.
    Útil para evitar problemas de compatibilidad en Python 3.14.
    """
    year, start_date, end_date = get_panel_date_range(request)
    return start_date.isoformat(), end_date.isoformat(), year


def fecha_a_iso(fecha):
    """
    Convierte cualquier fecha a string ISO format.
    Maneja: date, datetime, string, None.
    """
    if fecha is None:
        return None
    if isinstance(fecha, str):
        return fecha
    if isinstance(fecha, datetime):
        return fecha.date().isoformat()
    if isinstance(fecha, date):
        return fecha.isoformat()
    return str(fecha)


def regenerar_tablas_financieras(request):
    """Alias que usa get_panel_date_range y regenerate_financial_tables."""
    year, start_date, end_date = get_panel_date_range(request)
    regenerate_financial_tables(start_date, end_date)
    return start_date, end_date, year


def home(request):
    """Vista principal del panel con selector de año."""
    year, start_date, end_date = get_panel_date_range(request)
    context = {
        'panel_year': year,
        'panel_years': PANEL_YEAR_CHOICES,
        'panel_start_date': start_date,
        'panel_end_date': end_date,
    }
    return render(request, "boran_app/home.html", context)


def set_panel_year(request):
    """Vista para cambiar el año del panel (POST)."""
    if request.method == 'POST':
        year_value = request.POST.get('panel_year')
        try:
            year_int = int(year_value)
        except (TypeError, ValueError):
            year_int = _get_default_panel_year()
        if year_int not in PANEL_YEAR_CHOICES:
            year_int = _get_default_panel_year()
        request.session['panel_year'] = year_int
        next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('home')
        return redirect(next_url)
    return redirect('home')


# Alias para compatibilidad
def cambiar_anno_fiscal(request):
    """Alias de set_panel_year para compatibilidad con URLs existentes."""
    return set_panel_year(request)


def generar_balance_inicial_anno(request):
    """
    Genera el Balance Inicial para un año nuevo basándose en los saldos finales del año anterior.
    Los saldos finales se calculan como: débitos - créditos para cada cuenta.
    """
    from .models import BalanceInicial, ResumenDebito, ResumenCredito
    from .cod_cuentas_balance import balance_rows
    
    if request.method == 'POST':
        anno_destino = int(request.POST.get('anno_destino', 2026))
        anno_origen = anno_destino - 1
        
        # Verificar si ya existe balance inicial para el año destino
        existe = BalanceInicial.objects.filter(fecha__year=anno_destino).exists()
        if existe:
            messages.warning(request, f"Ya existe Balance Inicial para {anno_destino}. Elimínalo primero si quieres regenerarlo.")
            return redirect('generar_balance_inicial')
        
        # Regenerar tablas financieras del año origen para obtener saldos actualizados
        fecha_inicio_origen = date(anno_origen, 1, 1)
        fecha_fin_origen = date(anno_origen, 12, 31)
        
        regenerar_ventas_consulta(start_date=fecha_inicio_origen, end_date=fecha_fin_origen)
        poblar_movimientos_unificados_debito(start_date=fecha_inicio_origen, end_date=fecha_fin_origen)
        poblar_movimientos_unificados_credito(start_date=fecha_inicio_origen, end_date=fecha_fin_origen)
        regenerar_resumenes_credito_debito()
        
        # Obtener saldos
        debitos_dict = {d.cuenta_debito: float(d.total_debito) for d in ResumenDebito.objects.all()}
        creditos_dict = {c.cuenta_credito: float(c.total_credito) for c in ResumenCredito.objects.all()}
        
        # Fecha del balance inicial del nuevo año
        fecha_balance = date(anno_destino, 1, 1)
        
        # Crear registros de Balance Inicial para el nuevo año
        nuevos_registros = []
        cuentas_creadas = 0
        
        for fila in balance_rows:
            codigo = fila['codigo']
            
            # SOLO cuentas de Activo (1xxxxxx) y Pasivo (2xxxxxx)
            # NO incluir cuentas de Resultados (3xxxxxx)
            if not (1000000 <= codigo <= 2999999):
                continue
            
            debito = debitos_dict.get(codigo, 0)
            credito = creditos_dict.get(codigo, 0)
            
            # Solo crear si hay saldo
            if debito > 0 or credito > 0:
                # Calcular saldo neto
                saldo_deudor = max(debito - credito, 0)
                saldo_acreedor = max(credito - debito, 0)
                
                # Crear registro
                registro = BalanceInicial(
                    fecha=fecha_balance,
                    cuenta_debito=codigo,
                    debito=saldo_deudor,
                    cuenta_credito=codigo,
                    credito=saldo_acreedor,
                    comentario=f"Saldo inicial {anno_destino} (desde cierre {anno_origen})"
                )
                nuevos_registros.append(registro)
                cuentas_creadas += 1
        
        # Guardar todos los registros
        if nuevos_registros:
            BalanceInicial.objects.bulk_create(nuevos_registros)
            messages.success(request, f"✅ Balance Inicial {anno_destino} creado exitosamente con {cuentas_creadas} cuentas.")
        else:
            messages.warning(request, f"No se encontraron saldos para crear el Balance Inicial {anno_destino}.")
        
        return redirect('generar_balance_inicial')
    
    # GET: Mostrar formulario
    # Verificar qué años tienen balance inicial
    from django.db.models import Min, Max
    from django.db.models.functions import ExtractYear
    
    annos_con_balance = list(
        BalanceInicial.objects.annotate(anno=ExtractYear('fecha'))
        .values_list('anno', flat=True)
        .distinct()
        .order_by('anno')
    )
    
    # Contar registros por año
    conteo_por_anno = {}
    for anno in annos_con_balance:
        conteo_por_anno[anno] = BalanceInicial.objects.filter(fecha__year=anno).count()
    
    return render(request, 'boran_app/generar_balance_inicial.html', {
        'annos_con_balance': annos_con_balance,
        'conteo_por_anno': conteo_por_anno,
        'anno_sugerido': max(annos_con_balance) + 1 if annos_con_balance else 2025,
    })


def eliminar_balance_inicial_anno(request):
    """
    Elimina todos los registros de Balance Inicial de un año específico.
    """
    from .models import BalanceInicial
    
    if request.method == 'POST':
        anno = int(request.POST.get('anno', 0))
        if anno >= 2025:
            eliminados = BalanceInicial.objects.filter(fecha__year=anno).delete()[0]
            messages.success(request, f"✅ Se eliminaron {eliminados} registros del Balance Inicial {anno}.")
        else:
            messages.error(request, "Año inválido.")
    
    return redirect('generar_balance_inicial')

def listado_union_credito(request):
    total = MovimientoUnificadoCredito.objects.count()

    # <-- Aquí va la lógica para devolver JSON si se pide:
    if request.GET.get('ver') == 'json':
        data = list(
            MovimientoUnificadoCredito.objects.values(
                'fecha', 'cta_credito', 'monto_credito', 'texto_coment', 'tabla_origen'
            )
        )
        return JsonResponse(data, safe=False)

    # Si no llega ?ver=json, sigue mostrando la plantilla con el botón
    return render(request, 'boran_app/union_credito.html', {
        'total_registros': total
    })

def cargar_union_credito(request):
    """
    Esta vista se invoca cuando el usuario hace clic en “Cargar movimientos unificados”.
    -- Llama a poblar_movimientos_unificados() para insertar/actualizar.
    -- Luego redirige a listado_union y muestra un mensaje “Cargado con éxito”.
    """
    try:
        poblar_movimientos_unificados_credito()
        messages.success(request, f"Tabla 'movimiento_unificado' actualizada: ahora hay {MovimientoUnificadoCredito.objects.count()} registros.")
    except Exception as e:
        # Captura cualquier error y muestra mensaje de error
        messages.error(request, f"Error al poblar movimientos unificados: {e}")

    return redirect(reverse('listado_union_credito'))

def listado_union_debito(request):
    """
    Similar a listado_union, pero para la tabla de débitos.
    """
    total = MovimientoUnificadoDebito.objects.count()
    return render(request, 'boran_app/union_debito.html', {
        'total_registros': total
    })

def cargar_union_debito(request):
    """
    Dispara la carga para la tabla movemento_unificado_debito.
    """
    try:
        poblar_movimientos_unificados_debito()
        messages.success(request, f"Tabla 'movimiento_unificado_debito' actualizada: ahora hay {MovimientoUnificadoDebito.objects.count()} registros.")
    except Exception as e:
        messages.error(request, f"Error al poblar movimientos débito: {e}")

    return redirect(reverse('listado_union_debito'))

# ——————————————————————————————————————————————————————————————
#  VENTAS CONSULTA
# ————————————————————————————————————————————————————————
from django.shortcuts import redirect, render
from django.contrib.admin.views.decorators import staff_member_required
from .utils import regenerar_ventas_consulta

@staff_member_required
def regenerar_consulta_view(request):
    if request.method == "POST":
        regenerar_ventas_consulta()
        return render(request, "boran_app/confirmacion.html", {"mensaje": "✅ VentasConsulta regenerada correctamente."})

    return render(request, "boran_app/confirmacion.html", {"mensaje": None})   

# ——————————————————————————————————————————————————————————————
#  PAGINA INICIO
# ————————————————————————————————————————————————————————

@staff_member_required
def pagina_inicio(request):
    return render(request, "boran_app/inicio.html")

@staff_member_required
def regenerar_resumenes_view(request):
    mensaje = None
    if request.method == "POST":
        total_creditos, total_debitos = regenerar_resumenes_credito_debito()
        mensaje = f"✅ Regenerado: {total_creditos} créditos y {total_debitos} débitos."

    return render(request, "boran_app/resumen_regenerar.html", {"mensaje": mensaje})

# ——————————————————————————————————————————————————————————————
#  EXPORTAR A EXCEL SUMA CREDITOS SUMA DEBITOS A GESTION BORAN PARA BALANCE
# ————————————————————————————————————————————————————————


@staff_member_required
@csrf_exempt
def exportar_resumen_excel(request):
    if request.method == "POST":
        ruta_libro = r"C:\Users\tcort\OneDrive\BORAN\Django\Otros\GestionBoranX.xlsm"
        hoja_destino = "Datos"
        if not os.path.exists(ruta_libro):
            messages.error(request, f"❌ Archivo no encontrado: {ruta_libro}")
            return redirect('home')

        # --- Resúmenes y movimientos ---
        df_credito = pd.DataFrame(list(ResumenCredito.objects.values("cuenta_credito", "total_credito")))
        df_debito = pd.DataFrame(list(ResumenDebito.objects.values("cuenta_debito", "total_debito")))
        df_unif_credito = pd.DataFrame(list(MovimientoUnificadoCredito.objects.values("fecha", "cta_credito", "monto_credito")))
        df_unif_debito  = pd.DataFrame(list(MovimientoUnificadoDebito.objects.values("fecha", "cta_debito", "monto_debito")))

        filas_credito = len(df_credito)
        filas_debito = len(df_debito)
        filas_unif_credito = len(df_unif_credito)
        filas_unif_debito  = len(df_unif_debito)

        # --- Abrir Excel ---
        libro = openpyxl.load_workbook(ruta_libro, keep_vba=True)
        if hoja_destino not in libro.sheetnames:
            messages.error(request, f"❌ Hoja '{hoja_destino}' no encontrada en el libro.")
            return redirect('home')
        hoja = libro[hoja_destino]

        # Limpiar áreas destino
        for fila in hoja.iter_rows(min_row=2, min_col=1, max_col=3):   # A-C
            for celda in fila: celda.value = None
        for fila in hoja.iter_rows(min_row=2, min_col=5, max_col=7):   # E-G
            for celda in fila: celda.value = None
        for fila in hoja.iter_rows(min_row=2, min_col=10, max_col=13): # J-M
            for celda in fila: celda.value = None

        # Escribir movimientos
        for i, row in df_unif_debito.iterrows():
            hoja[f"A{i+2}"] = row["fecha"]
            hoja[f"B{i+2}"] = row["cta_debito"]
            hoja[f"C{i+2}"] = row["monto_debito"]
        for i, row in df_unif_credito.iterrows():
            hoja[f"E{i+2}"] = row["fecha"]
            hoja[f"F{i+2}"] = row["cta_credito"]
            hoja[f"G{i+2}"] = row["monto_credito"]

        # Escribir resumenes
        max_filas = max(filas_credito, filas_debito)
        for i in range(max_filas):
            hoja[f"J{i+2}"] = df_debito.iloc[i]["cuenta_debito"] if i < filas_debito else ""
            hoja[f"K{i+2}"] = df_debito.iloc[i]["total_debito"] if i < filas_debito else ""
            hoja[f"L{i+2}"] = df_credito.iloc[i]["cuenta_credito"] if i < filas_credito else ""
            hoja[f"M{i+2}"] = df_credito.iloc[i]["total_credito"] if i < filas_credito else ""

        libro.save(ruta_libro)

        # Un solo mensaje, sólo HTML seguro
        messages.success(
            request,
            (
                "✅ Exportación completada.<br>"
                f"📊 Créditos exportados (resumen): {filas_credito} <br>"
                f"📊 Débitos exportados (resumen): {filas_debito} <br>"
                f"📤 Movimientos crédito: {filas_unif_credito} <br>"
                f"📤 Movimientos débito: {filas_unif_debito}"
            )
        )
        return redirect('home')

    # Redirige siempre a home si la petición no es POST
    return redirect('home')

# ——————————————————————————————————————————————————————————————
#  EXPORTAR A EXCEL TODOS CREDITOS Y TODOS DEBITOS A GESTION BORANGORA BALANCE
# ————————————————————————————————————————————————————————

from boran_app.scripts.export_a_excel import main as export_a_excel_script
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.shortcuts import redirect

@require_POST
def export_a_excel_view(request):
    msg = exportar_resumen_excel()
    if "Error" in msg or "error" in msg or "🧨" in msg:
        messages.error(request, msg)
    else:
        messages.success(request, msg)
    return redirect('home')

 # ——————————————————————————————————————————————————————————————
# VISTAS EN HOME IMPORTACIONES
# ————————————————————————————————————————————————————————   

from django.shortcuts import redirect, render
from django.contrib import messages
from django.views.decorators.http import require_POST
from boran_app.scripts.import_ajuste_inventario import main as import_ajuste_inventario_script
from boran_app.scripts.import_asientos_contables import main as import_asientos_contables_script
from boran_app.scripts.import_catalogo import main as import_catalogo_script
from boran_app.scripts.import_balance_inicial import main as import_balance_inicial_script
from boran_app.scripts.import_entrada_productos import main as import_entrada_productos_script
from boran_app.scripts.import_inventario_inicial import main as import_inventario_inicial_script
from boran_app.scripts.import_sueldos import main as import_sueldos_script
from boran_app.scripts.import_ventas import main as import_ventas_script
from boran_app.scripts.import_envios import main as import_envios_script
from boran_app.scripts.import_otros_gastos import main as import_otros_gastos_script
from boran_app.scripts.import_envios_a_tiendas import main as import_envios_a_tiendas_script

@require_POST
def import_ajuste_inventario(request):
    try:
        mensaje = import_ajuste_inventario_script()
        messages.success(request, mensaje)
    except Exception as e:
        messages.error(request, f"Error al importar Ajuste de Inventario: {str(e)}")
    return redirect('home')

@require_POST
def import_asientos_contables(request):
    try:
        msg = import_asientos_contables_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Asientos Contables: {str(e)}")
    return redirect('home')

@require_POST
def import_catalogo(request):
    try:
        resultado = import_catalogo_script()
        if resultado and "no hay" in resultado.lower():
            messages.info(request, resultado)
        else:
            messages.success(request, resultado)
    except Exception as e:
        messages.error(request, f"Error al importar Catálogo: {str(e)}")
    return redirect('home')

@require_POST
def import_balance_inicial(request):
    try:
        msg = import_balance_inicial_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Balance Inicial: {str(e)}")
    return redirect('home')

@require_POST
def import_entrada_productos(request):
    try:
        msg = import_entrada_productos_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Entradas de Productos: {str(e)}")
    return redirect('home')

@require_POST
def import_inventario_inicial(request):
    try:
        msg = import_inventario_inicial_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Inventario Inicial: {str(e)}")
    return redirect('home')

@require_POST
def import_sueldos(request):
    try:
        msg = import_sueldos_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Sueldos: {str(e)}")
    return redirect('home')

@require_POST
def import_ventas(request):
    try:
        msg = import_ventas_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Ventas: {str(e)}")
    return redirect('home')

@require_POST
def import_envios(request):
    try:
        mensaje = import_envios_script()
        messages.success(request, mensaje)
    except Exception as e:
        messages.error(request, f"Error al importar Envíos: {str(e)}")
    return redirect('home')

@require_POST
def import_otros_gastos(request):
    try:
        msg = import_otros_gastos_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Otros Gastos: {str(e)}")
    return redirect('home')

@require_POST
def import_envios_a_tiendas(request):
    try:
        msg = import_envios_a_tiendas_script()
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Error al importar Envios a Tiendas: {str(e)}")
    return redirect('home')

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect

@login_required
def procesar_todo(request):
    if request.method == 'POST':
        try:
            # Ejecutar en orden las 4 funciones utilitarias
            poblar_movimientos_unificados_credito()
            poblar_movimientos_unificados_debito()
            regenerar_resumenes_credito_debito()
            regenerar_ventas_consulta()
            messages.success(request, "¡Todos los procesos ejecutados correctamente en orden:! (VentasConsulta, Unión Créditos, Unión Débitos y Resúmenes)")
        except Exception as e:
            messages.error(request, f"Error al ejecutar procesamiento total: {e}")
    return redirect('home')


# ——————————————————————————————————————————————————————————————
# VISTAS EN HOME PROCESAR DATOS
# ————————————————————————————————————————————————————————   

from django.shortcuts import redirect
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt

# Importa tus funciones utilitarias de procesamiento aquí
from .utils import (
    regenerar_ventas_consulta,
    poblar_movimientos_unificados_credito,
    poblar_movimientos_unificados_debito,
    regenerar_resumenes_credito_debito, 
)

from django.contrib.auth.decorators import login_required

@login_required
def procesar_ventas_consulta(request):
    if request.method == 'POST':
        regenerar_ventas_consulta()
        messages.success(request, "¡Tabla VentasConsulta regenerada!")
    return redirect('home')

@login_required
def procesar_union_credito(request):
    if request.method == 'POST':
        poblar_movimientos_unificados_credito()
        messages.success(request, "¡Unión de Créditos procesada con éxito!")
    return redirect('home')

@login_required
def procesar_union_debito(request):
    if request.method == 'POST':
        poblar_movimientos_unificados_debito()
        messages.success(request, "¡Unión de Débitos procesada con éxito!")
    return redirect('home')

@login_required
def procesar_resumenes(request):
    if request.method == 'POST':
        regenerar_resumenes_credito_debito()
        messages.success(request, "¡Resúmenes de crédito y débito regenerados!")
    return redirect('home')

# ——————————————————————————————————————————————————————————————
# VISTAS EN HOME INVENTARIO
# ————————————————————————————————————————————————————————   

from django.contrib import messages
from django.core.management import call_command
import io
import re

@login_required
def procesar_inventario(request):
    if request.method == 'POST':
        out = io.StringIO()
        try:
            call_command('inventario', stdout=out)
            resultado = out.getvalue()
            # Limpiar ANSI:
            ansi_escape = re.compile(r'\x1B[@-_][0-?]*[ -/]*[@-~]')
            resultado_limpio = ansi_escape.sub('', resultado)
            messages.success(request, f"<pre>{resultado_limpio}</pre>")
        except Exception as e:
            messages.error(request, f"Error al procesar inventario: {e}")
    return redirect('home')

from django.http import HttpResponse
from openpyxl import Workbook
from .models import Catalogo, InventarioInicial, EntradaProductos, Envios, Ventas, AjusteInventario
from django.db import models
from django.urls import reverse

from django.http import HttpResponse
from openpyxl import Workbook
from .models import Catalogo, InventarioInicial, EntradaProductos, Envios, Ventas, AjusteInventario
from django.db import models

def exportar_inventario_actual(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario Actual"

    # Cabecera con columnas calculadas
    ws.append([
        'SKU', 'Categoría', 'Producto',
        'Stock', 'Bodega', 'Ingresos', 'Envios', 'Ventas', 'Ajustes',
        'En Oficina', 'En Bodega', 'Ajuste Ventas', 'Total'
    ])

    for obj in Catalogo.objects.all():
        ini = InventarioInicial.objects.filter(sku=obj.sku).first()
        inicial = ini.stock if ini else 0
        bodega = ini.bodega if ini else 0

        ingresos = EntradaProductos.objects.filter(
            sku__sku=obj.sku
        ).aggregate(total=models.Sum('cantidad_ingresada'))['total'] or 0

        envios = Envios.objects.filter(
            sku__sku=obj.sku
        ).aggregate(total=models.Sum('cantidad'))['total'] or 0

        ventas = Ventas.objects.filter(
            sku__sku=obj.sku
        ).aggregate(total=models.Sum('cantidad'))['total'] or 0

        ajustes = AjusteInventario.objects.filter(
            sku__sku=obj.sku
        ).aggregate(total=models.Sum('cantidad'))['total'] or 0

        # MISMA LÓGICA QUE LA VISTA inventario_actual
        en_oficina = 0
        en_bodega = inicial + ingresos - ventas      # igual que en la vista
        ajuste_ventas = 0                            # de momento fijo
        total = en_oficina + en_bodega - ajustes     # igual que en la vista

        ws.append([
            obj.sku,
            obj.categoria,
            obj.producto,
            inicial,
            bodega,
            ingresos,
            envios,
            ventas,
            ajustes,
            en_oficina,
            en_bodega,
            ajuste_ventas,
            total,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=InventarioActual.xlsx'
    wb.save(response)
    return response


# ——————————————————————————————————————————————————————————————
# TABLAS PARA INVENTARIO
# ——————————————————————————————————————————————————————————————


from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import reverse
from .models import Catalogo, InventarioInicial, EntradaProductos, Envios, Ventas, AjusteInventario
from django.db import models

def inventario_actual(request):
    q = request.GET.get('q', '').strip()
    sort = request.GET.get('sort', 'sku')
    direction = request.GET.get('dir', 'asc')

    columnas_ordenables = {
        'sku': 'sku',
        'categoria': 'categoria',
        'producto': 'producto',
        'inicial': None,
        'bodega': None,
        'ingresos': None,
        'envios': None,
        'ventas': None,
        'ajustes': None,
        'en_oficina': None,
        'en_bodega': None,
        #'ajuste_ventas': None,
        'total': None,
    }

    todos = Catalogo.objects.all()
    if q:
        todos = todos.filter(
            models.Q(sku__icontains=q) |
            models.Q(categoria__icontains=q) |
            models.Q(producto__icontains=q)
        )

    if columnas_ordenables.get(sort):
        orden = columnas_ordenables[sort]
        if direction == 'desc':
            orden = '-' + orden
        todos = todos.order_by(orden)

        # ---- Prepara ventas por SKU para ajuste ----
    ventas_por_sku = {}
    for obj in todos:
        ventas = Ventas.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad'))['total'] or 0
        ventas_por_sku[obj.sku] = ventas

    productos = []
    for obj in todos:
        ini = InventarioInicial.objects.filter(sku=obj.sku).first()
        inicial = ini.stock if ini else 0
        bodega = ini.bodega if ini else 0
        ingresos = EntradaProductos.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad_ingresada'))['total'] or 0
        envios = Envios.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad'))['total'] or 0
        ventas = ventas_por_sku.get(obj.sku, 0)
        ajustes = AjusteInventario.objects.filter(sku__sku=obj.sku).aggregate(total=models.Sum('cantidad'))['total'] or 0

        en_oficina = inicial + ingresos - envios                  # o tu fórmula futura
        en_bodega = inicial + envios - ventas - ajustes
        total = en_oficina + en_bodega

        productos.append({
            'sku': obj.sku,
            'categoria': obj.categoria,
            'producto': obj.producto,
            'inicial': inicial,
            'bodega': bodega,
            'ingresos': ingresos,
            'envios': envios,
            'ventas': ventas,
            'ajustes': ajustes,
            'en_oficina': en_oficina,
            'en_bodega': en_bodega,
            #'ajuste_ventas': ajuste_ventas,
            'total': total,
        })


    # Ordena si corresponde
    if columnas_ordenables.get(sort) is None and productos and sort in productos[0]:
        reverse_sort = (direction == 'desc')
        productos = sorted(productos, key=lambda x: x[sort], reverse=reverse_sort)

    # Totales
    global_total_inicial = sum(p['inicial'] for p in productos)
    global_total_bodega = sum(p['bodega'] for p in productos)
    global_total_ingresos = sum(p['ingresos'] for p in productos)
    global_total_envios = sum(p['envios'] for p in productos)
    global_total_ventas = sum(p['ventas'] for p in productos)
    global_total_ajustes = sum(p['ajustes'] for p in productos)
    global_total_en_oficina = sum(p['en_oficina'] for p in productos)
    global_total_en_bodega = sum(p['en_bodega'] for p in productos)
    #global_total_ajuste_ventas = sum(p['ajuste_ventas'] for p in productos)
    global_total_total = sum(p['total'] for p in productos)

    paginator = Paginator(productos, 10)
    page = request.GET.get('page')
    try:
        productos_page = paginator.page(page)
    except PageNotAnInteger:
        productos_page = paginator.page(1)
    except EmptyPage:
        productos_page = paginator.page(paginator.num_pages)

    total_inicial = sum(p['inicial'] for p in productos_page.object_list)
    total_bodega = sum(p['bodega'] for p in productos_page.object_list)
    total_ingresos = sum(p['ingresos'] for p in productos_page.object_list)
    total_envios = sum(p['envios'] for p in productos_page.object_list)
    total_ventas = sum(p['ventas'] for p in productos_page.object_list)
    total_ajustes = sum(p['ajustes'] for p in productos_page.object_list)
    total_en_oficina = sum(p['en_oficina'] for p in productos_page.object_list)
    total_en_bodega = sum(p['en_bodega'] for p in productos_page.object_list)
    #total_ajuste_ventas = sum(p['ajuste_ventas'] for p in productos_page.object_list)
    total_total = sum(p['total'] for p in productos_page.object_list)

    url_excel = reverse('exportar_inventario_actual')
    return render(request, 'boran_app/inventario.html', {
        'productos_page': productos_page,
        'q': q,
        'url_excel': url_excel,
        'sort': sort,
        'direction': direction,
        'total_inicial': total_inicial,
        'total_bodega': total_bodega,
        'total_ingresos': total_ingresos,
        'total_envios': total_envios,
        'total_ventas': total_ventas,
        'total_ajustes': total_ajustes,
        'total_en_oficina': total_en_oficina,
        'total_en_bodega': total_en_bodega,
        #'total_ajuste_ventas': total_ajuste_ventas,
        'total_total': total_total,
        'global_total_inicial': global_total_inicial,
        'global_total_bodega': global_total_bodega,
        'global_total_ingresos': global_total_ingresos,
        'global_total_envios': global_total_envios,
        'global_total_ventas': global_total_ventas,
        'global_total_ajustes': global_total_ajustes,
        'global_total_en_oficina': global_total_en_oficina,
        'global_total_en_bodega': global_total_en_bodega,
        #'global_total_ajuste_ventas': global_total_ajuste_ventas,
        'global_total_total': global_total_total,
    })



# ——————————————————————————————————————————————————————————————
# BALANCE
# ——————————————————————————————————————————————————————————————

from datetime import date  # <-- Importa SOLO la clase date
from .utils import (
    regenerar_ventas_consulta,
    poblar_movimientos_unificados_debito,
    poblar_movimientos_unificados_credito,
    regenerar_resumenes_credito_debito,
    
)

from django.shortcuts import render
from django.http import HttpResponse
from .models import ResumenDebito, ResumenCredito
import pandas as pd
from .cod_cuentas_balance import balance_rows

def intdot(val):
    """Formatea números para mostrar miles con punto."""
    try:
        val_float = float(val)
        val_int = int(round(val_float))
        return f"{val_int:,}".replace(",", ".")
    except Exception:
        return ""

def balance_view(request):

    # Obtener fechas del año fiscal y regenerar tablas
    fecha_inicio, fecha_fin, anno_fiscal = regenerar_tablas_financieras(request)

    debitos_dict = {d.cuenta_debito: float(d.total_debito) for d in ResumenDebito.objects.all()}
    creditos_dict = {c.cuenta_credito: float(c.total_credito) for c in ResumenCredito.objects.all()}

    matriz_balance = []
    fecha_corte = f"01/01/{anno_fiscal} - 31/12/{anno_fiscal}"
    total_debito = total_credito = total_saldo_deudor = total_saldo_acreedor = 0
    total_activo = total_pasivo = total_perdidas = total_ganancias = 0

    for fila in balance_rows:
        codigo = fila['codigo']
        nombre = fila['nombre']
        debito = debitos_dict.get(codigo, 0)
        credito = creditos_dict.get(codigo, 0)
        saldo_deudor = saldo_acreedor = activo = pasivo = perdidas = ganancias = 0

        if 1010100 <= codigo <= 2040000:
            saldo_deudor = debito - credito if debito > credito else 0
            saldo_acreedor = credito - debito if credito > debito else 0
            activo = saldo_deudor
            pasivo = saldo_acreedor
        elif 3010100 <= codigo <= 3030300:
            saldo_deudor = debito - credito if debito > credito else 0
            saldo_acreedor = credito - debito if credito > debito else 0
            perdidas = saldo_deudor
            ganancias = saldo_acreedor

        matriz_balance.append({
            'codigo': codigo,
            'nombre': nombre,
            'debito': intdot(debito),
            'credito': intdot(credito),
            'saldo_deudor': intdot(saldo_deudor),
            'saldo_acreedor': intdot(saldo_acreedor),
            'activo': intdot(activo),
            'pasivo': intdot(pasivo),
            'perdidas': intdot(perdidas),
            'ganancias': intdot(ganancias)
        })

        # Acumula totales
        total_debito += debito
        total_credito += credito
        total_saldo_deudor += saldo_deudor
        total_saldo_acreedor += saldo_acreedor
        total_activo += activo
        total_pasivo += pasivo
        total_perdidas += perdidas
        total_ganancias += ganancias

    # -- Agrega fila: Utilidad (pérdida) del Ejercicio
    def resta_positiva(a, b):
        return max(0, a - b)

    utilidad_activo = resta_positiva(total_pasivo, total_activo)
    utilidad_pasivo = resta_positiva(total_activo, total_pasivo)
    utilidad_perdidas = resta_positiva(total_ganancias, total_perdidas)
    utilidad_ganancias = resta_positiva(total_perdidas, total_ganancias)

    utilidad = {
        'debito': '',
        'credito': '',
        'saldo_deudor': '',
        'saldo_acreedor': '',
        'activo': intdot(utilidad_activo),
        'pasivo': intdot(utilidad_pasivo),
        'perdidas': intdot(utilidad_perdidas),
        'ganancias': intdot(utilidad_ganancias)
    }

    # Flags para pintar en rojo
    utilidad_pasivo_rojo = total_activo < total_pasivo
    utilidad_perdidas_rojo = total_perdidas > total_ganancias

    # -- Agrega fila: SUMAS TOTALES (Totales + Utilidad)
    sumas_totales = {
        'debito': intdot(total_debito),
        'credito': intdot(total_credito),
        'saldo_deudor': intdot(total_saldo_deudor),
        'saldo_acreedor': intdot(total_saldo_acreedor),
        'activo': intdot(total_activo + utilidad_activo),
        'pasivo': intdot(total_pasivo + utilidad_pasivo),
        'perdidas': intdot(total_perdidas + utilidad_perdidas),
        'ganancias': intdot(total_ganancias + utilidad_ganancias)
    }

    totales = {
        'debito': intdot(total_debito),
        'credito': intdot(total_credito),
        'saldo_deudor': intdot(total_saldo_deudor),
        'saldo_acreedor': intdot(total_saldo_acreedor),
        'activo': intdot(total_activo),
        'pasivo': intdot(total_pasivo),
        'perdidas': intdot(total_perdidas),
        'ganancias': intdot(total_ganancias)
    }

    # Exportar a Excel si es solicitado
    if request.GET.get("export") == "excel":
        df = pd.DataFrame([
            {**f, **{k: v.replace('.', '') for k, v in f.items() if k not in ('codigo', 'nombre')}}
            for f in matriz_balance
        ])
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="balance.xlsx"'
        df.to_excel(response, index=False)
        return response

    return render(request, "boran_app/balance.html", {
        'matriz_balance': matriz_balance,
        'totales': totales,
        'utilidad': utilidad,
        'sumas_totales': sumas_totales,
        'utilidad_pasivo_rojo': utilidad_pasivo_rojo,
        'utilidad_perdidas_rojo': utilidad_perdidas_rojo,
        'fecha_corte': fecha_corte,  # <--- Aquí pasas la fecha de hoy
    })

# ——————————————————————————————————————————————————————————————
# RESUMEN BALANCE
# ——————————————————————————————————————————————————————————————

import json
from django.shortcuts import render

def resumen_balance_view(request):

    # Obtener fechas del año fiscal y regenerar tablas
    fecha_inicio, fecha_fin, anno_fiscal = regenerar_tablas_financieras(request)

    # 1. Prepara los datos "matriz_dict"
    # Debe ser: {'A:1010100': 123, 'P:1010100': 0, ...}
    # Puedes obtenerlo de tu lógica existente para balance_view
    # Aquí te doy un ejemplo muy simple:
    matriz_dict = {}
    # Supón que tienes lista de filas balance_rows y los valores de cada columna
    # Puedes reutilizar tu código de balance_view:
    from .models import ResumenDebito, ResumenCredito
    from .cod_cuentas_balance import balance_rows

    debitos_dict = {d.cuenta_debito: float(d.total_debito) for d in ResumenDebito.objects.all()}
    creditos_dict = {c.cuenta_credito: float(c.total_credito) for c in ResumenCredito.objects.all()}
    for fila in balance_rows:
        codigo = str(fila['codigo'])
        debito = debitos_dict.get(int(codigo), 0)
        credito = creditos_dict.get(int(codigo), 0)
        # Calcula las columnas igual que tu balance_view
        saldo_deudor = debito - credito if debito > credito else 0
        saldo_acreedor = credito - debito if credito > debito else 0
        # Según rango, asigna a A, P, Pe, G
        if 1010100 <= int(codigo) <= 2040000:
            matriz_dict[f'A:{codigo}'] = saldo_deudor
            matriz_dict[f'P:{codigo}'] = saldo_acreedor
        elif 3010100 <= int(codigo) <= 3010300:
            matriz_dict[f'Pe:{codigo}'] = saldo_deudor
            matriz_dict[f'G:{codigo}'] = saldo_acreedor

    # 2. Renderiza el template y pasa matriz_js
    return render(request, "boran_app/resumen_balance.html", {
        'matriz_js': json.dumps(matriz_dict),
    })

# ——————————————————————————————————————————————————————————————
# BALANCE SEGUN FECHA
# ——————————————————————————————————————————————————————————————

from django.shortcuts import render
from django.http import HttpResponse
from .models import MovimientoUnificadoCredito, MovimientoUnificadoDebito
from .cod_cuentas_balance import balance_rows
from datetime import datetime
import pandas as pd

def balance_segun_fecha_view(request):
    """
    Genera el balance según fecha de corte, respetando el año fiscal seleccionado.
    Solo procesa datos del año fiscal actual.
    """
    # Obtener año fiscal de la sesión
    anno_fiscal = get_panel_year(request)
    fecha_inicio_anno = date(anno_fiscal, 1, 1)
    
    # Obtener fecha de corte del parámetro GET
    fecha_corte = request.GET.get('fecha_corte')
    fecha_corte_dt = None
    fecha_corte_str = ''

    if fecha_corte:
        try:
            fecha_corte_dt = datetime.strptime(fecha_corte, "%Y-%m-%d").date()
            fecha_corte_str = fecha_corte_dt.strftime("%Y-%m-%d")
        except Exception:
            fecha_corte_dt = None
            fecha_corte_str = ''

    if not fecha_corte_dt:
        # Mostrar formulario con el año fiscal actual
        return render(request, "boran_app/balance_segun_fecha.html", {
            'panel_year': anno_fiscal,
        })
    
    # Regenerar tablas financieras con el rango del año fiscal hasta la fecha de corte
    regenerar_ventas_consulta(start_date=fecha_inicio_anno, end_date=fecha_corte_dt)
    poblar_movimientos_unificados_debito(start_date=fecha_inicio_anno, end_date=fecha_corte_dt)
    poblar_movimientos_unificados_credito(start_date=fecha_inicio_anno, end_date=fecha_corte_dt)
    regenerar_resumenes_credito_debito()

    # Filtrar movimientos desde el inicio del año fiscal hasta la fecha de corte
    # Usar fecha_a_iso para compatibilidad con Python 3.14
    debitos = MovimientoUnificadoDebito.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio_anno),
        fecha__lte=fecha_a_iso(fecha_corte_dt)
    )
    creditos = MovimientoUnificadoCredito.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio_anno),
        fecha__lte=fecha_a_iso(fecha_corte_dt)
    )

    debitos_dict = {}
    for d in debitos:
        debitos_dict[d.cta_debito] = debitos_dict.get(d.cta_debito, 0) + float(d.monto_debito)
    creditos_dict = {}
    for c in creditos:
        creditos_dict[c.cta_credito] = creditos_dict.get(c.cta_credito, 0) + float(c.monto_credito)

    matriz_balance = []
    total_debito = total_credito = total_saldo_deudor = total_saldo_acreedor = 0
    total_activo = total_pasivo = total_perdidas = total_ganancias = 0

    def intdot(val):
        try:
            val_float = float(val)
            val_int = int(round(val_float))
            return f"{val_int:,}".replace(",", ".")
        except Exception:
            return ""

    for fila in balance_rows:
        codigo = fila['codigo']
        nombre = fila['nombre']
        debito = debitos_dict.get(codigo, 0)
        credito = creditos_dict.get(codigo, 0)
        saldo_deudor = saldo_acreedor = activo = pasivo = perdidas = ganancias = 0

        if 1010100 <= codigo <= 2040000:
            saldo_deudor = debito - credito if debito > credito else 0
            saldo_acreedor = credito - debito if credito > debito else 0
            activo = saldo_deudor
            pasivo = saldo_acreedor
        elif 3010100 <= codigo <= 3030300:
            saldo_deudor = debito - credito if debito > credito else 0
            saldo_acreedor = credito - debito if credito > debito else 0
            perdidas = saldo_deudor
            ganancias = saldo_acreedor

        matriz_balance.append({
            'codigo': codigo,
            'nombre': nombre,
            'debito': intdot(debito),
            'credito': intdot(credito),
            'saldo_deudor': intdot(saldo_deudor),
            'saldo_acreedor': intdot(saldo_acreedor),
            'activo': intdot(activo),
            'pasivo': intdot(pasivo),
            'perdidas': intdot(perdidas),
            'ganancias': intdot(ganancias)
        })

        total_debito += debito
        total_credito += credito
        total_saldo_deudor += saldo_deudor
        total_saldo_acreedor += saldo_acreedor
        total_activo += activo
        total_pasivo += pasivo
        total_perdidas += perdidas
        total_ganancias += ganancias

    def resta_positiva(a, b):
        return max(0, a - b)

    utilidad_activo = resta_positiva(total_pasivo, total_activo)
    utilidad_pasivo = resta_positiva(total_activo, total_pasivo)
    utilidad_perdidas = resta_positiva(total_ganancias, total_perdidas)
    utilidad_ganancias = resta_positiva(total_perdidas, total_ganancias)

    utilidad = {
        'debito': '',
        'credito': '',
        'saldo_deudor': '',
        'saldo_acreedor': '',
        'activo': intdot(utilidad_activo),
        'pasivo': intdot(utilidad_pasivo),
        'perdidas': intdot(utilidad_perdidas),
        'ganancias': intdot(utilidad_ganancias)
    }

    utilidad_pasivo_rojo = total_activo < total_pasivo
    utilidad_perdidas_rojo = total_perdidas > total_ganancias

    sumas_totales = {
        'debito': intdot(total_debito),
        'credito': intdot(total_credito),
        'saldo_deudor': intdot(total_saldo_deudor),
        'saldo_acreedor': intdot(total_saldo_acreedor),
        'activo': intdot(total_activo + utilidad_activo),
        'pasivo': intdot(total_pasivo + utilidad_pasivo),
        'perdidas': intdot(total_perdidas + utilidad_perdidas),
        'ganancias': intdot(total_ganancias + utilidad_ganancias)
    }

    totales = {
        'debito': intdot(total_debito),
        'credito': intdot(total_credito),
        'saldo_deudor': intdot(total_saldo_deudor),
        'saldo_acreedor': intdot(total_saldo_acreedor),
        'activo': intdot(total_activo),
        'pasivo': intdot(total_pasivo),
        'perdidas': intdot(total_perdidas),
        'ganancias': intdot(total_ganancias)
    }

    # Exportar a Excel si es solicitado
    if request.GET.get("export") == "excel":
        df = pd.DataFrame([
            {**f, **{k: v.replace('.', '') for k, v in f.items() if k not in ('codigo', 'nombre')}}
            for f in matriz_balance
        ])
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="balance_segun_fecha_{fecha_corte_str}.xlsx"'
        df.to_excel(response, index=False)
        return response

    return render(request, "boran_app/balance.html", {
        'matriz_balance': matriz_balance,
        'totales': totales,
        'utilidad': utilidad,
        'sumas_totales': sumas_totales,
        'utilidad_pasivo_rojo': utilidad_pasivo_rojo,
        'utilidad_perdidas_rojo': utilidad_perdidas_rojo,
        'fecha_corte': fecha_corte_str,
    })


# ————————————————————————————————————————————————————————
# RESUMEN BALANCE SEGÚN FECHA
# ————————————————————————————————————————————————————————
from django.shortcuts import render
import json
from .models import MovimientoUnificadoCredito, MovimientoUnificadoDebito
from .cod_cuentas_balance import balance_rows

def resumen_balance_segun_fecha_view(request):
    """
    Genera el resumen de balance según fecha de corte, respetando el año fiscal seleccionado.
    """
    # Obtener año fiscal de la sesión
    anno_fiscal = get_panel_year(request)
    fecha_inicio_anno = date(anno_fiscal, 1, 1)
    
    fecha_corte = request.GET.get('fecha_corte')
    matriz_dict = {}
    fecha_corte_dt = None

    if fecha_corte:
        from datetime import datetime
        try:
            fecha_corte_dt = datetime.strptime(fecha_corte, "%Y-%m-%d").date()
        except ValueError:
            fecha_corte_dt = None
    
    if not fecha_corte_dt:
        return render(request, "boran_app/resumen_balance_segun_fecha.html", {
            'matriz_js': json.dumps(matriz_dict),
            'fecha_corte': fecha_corte,
            'panel_year': anno_fiscal,
        })
    
    # Regenerar tablas financieras con el rango del año fiscal hasta la fecha de corte
    regenerar_ventas_consulta(start_date=fecha_inicio_anno, end_date=fecha_corte_dt)
    poblar_movimientos_unificados_debito(start_date=fecha_inicio_anno, end_date=fecha_corte_dt)
    poblar_movimientos_unificados_credito(start_date=fecha_inicio_anno, end_date=fecha_corte_dt)
    regenerar_resumenes_credito_debito()

    # Filtrar movimientos desde el inicio del año fiscal hasta la fecha de corte
    # Usar fecha_a_iso para compatibilidad con Python 3.14
    debitos = MovimientoUnificadoDebito.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio_anno),
        fecha__lte=fecha_a_iso(fecha_corte_dt)
    )
    creditos = MovimientoUnificadoCredito.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio_anno),
        fecha__lte=fecha_a_iso(fecha_corte_dt)
    )
    
    debitos_dict = {}
    for d in debitos:
        debitos_dict[d.cta_debito] = debitos_dict.get(d.cta_debito, 0) + float(d.monto_debito)
    creditos_dict = {}
    for c in creditos:
        creditos_dict[c.cta_credito] = creditos_dict.get(c.cta_credito, 0) + float(c.monto_credito)
    for fila in balance_rows:
        codigo = str(fila['codigo'])
        debito = debitos_dict.get(int(codigo), 0)
        credito = creditos_dict.get(int(codigo), 0)
        saldo_deudor = debito - credito if debito > credito else 0
        saldo_acreedor = credito - debito if credito > debito else 0
        if 1010100 <= int(codigo) <= 2040000:
            matriz_dict[f'A:{codigo}'] = saldo_deudor
            matriz_dict[f'P:{codigo}'] = saldo_acreedor
        elif 3010100 <= int(codigo) <= 3030300:
            matriz_dict[f'Pe:{codigo}'] = saldo_deudor
            matriz_dict[f'G:{codigo}'] = saldo_acreedor

    return render(request, "boran_app/resumen_balance_segun_fecha.html", {
        'matriz_js': json.dumps(matriz_dict),
        'fecha_corte': fecha_corte,
        'panel_year': anno_fiscal,
    })


# ——————————————————————————————————————————————————————————————
# RESUMEN FINANCIERO - VIEW
# ——————————————————————————————————————————————————————————————

import json
from django.shortcuts import render

def resumen_financiero(request):
    # Obtener fechas del año fiscal y regenerar tablas
    fecha_inicio, fecha_fin, anno_fiscal = regenerar_tablas_financieras(request)

    matriz_dict = {}
    from .models import ResumenDebito, ResumenCredito
    from .cod_cuentas_balance import balance_rows   # o tu archivo resumen_financiero_rows si tienes uno diferente

    debitos_dict = {d.cuenta_debito: float(d.total_debito) for d in ResumenDebito.objects.all()}
    creditos_dict = {c.cuenta_credito: float(c.total_credito) for c in ResumenCredito.objects.all()}

    for fila in balance_rows:   # Si tienes otro set de rows para el financiero, cámbialo aquí
        codigo = str(fila['codigo'])
        debito = debitos_dict.get(int(codigo), 0)
        credito = creditos_dict.get(int(codigo), 0)
        saldo_deudor = debito - credito if debito > credito else 0
        saldo_acreedor = credito - debito if credito > debito else 0
        if 1010100 <= int(codigo) <= 2040000:
            matriz_dict[f'A:{codigo}'] = saldo_deudor
            matriz_dict[f'P:{codigo}'] = saldo_acreedor
        elif 3010100 <= int(codigo) <= 3030300:
            matriz_dict[f'Pe:{codigo}'] = saldo_deudor
            matriz_dict[f'G:{codigo}'] = saldo_acreedor

    return render(request, "boran_app/resumen_financiero.html", {
        'matriz_js': json.dumps(matriz_dict),
    })


# ——————————————————————————————————————————————————————————————
# RESUMEN FINANCIERO SEGUN FECHA CORTE- VIEW
# ——————————————————————————————————————————————————————————————

from django.shortcuts import render
import json
from .models import MovimientoUnificadoCredito, MovimientoUnificadoDebito
from .cod_cuentas_balance import balance_rows
from datetime import datetime

def resumen_financiero_segun_fecha_view(request):
    """
    Genera el resumen financiero según fecha de corte, respetando el año fiscal seleccionado.
    """
    # Obtener año fiscal de la sesión
    anno_fiscal = get_panel_year(request)
    fecha_inicio_anno = date(anno_fiscal, 1, 1)
    
    fecha_corte = request.GET.get('fecha_corte')
    matriz_dict = {}
    fecha_corte_dt = None

    if fecha_corte:
        try:
            # Permite ambos formatos de fecha: dd-mm-yyyy y yyyy-mm-dd
            try:
                fecha_corte_dt = datetime.strptime(fecha_corte, "%d-%m-%Y").date()
            except ValueError:
                fecha_corte_dt = datetime.strptime(fecha_corte, "%Y-%m-%d").date()
        except Exception:
            fecha_corte_dt = None
    
    if not fecha_corte_dt:
        return render(request, "boran_app/resumen_financiero_segun_fecha.html", {
            'matriz_js': json.dumps(matriz_dict),
            'fecha_corte': fecha_corte,
            'panel_year': anno_fiscal,
        })
    
    # Regenerar tablas financieras con el rango del año fiscal hasta la fecha de corte
    regenerar_ventas_consulta(start_date=fecha_inicio_anno, end_date=fecha_corte_dt)
    poblar_movimientos_unificados_debito(start_date=fecha_inicio_anno, end_date=fecha_corte_dt)
    poblar_movimientos_unificados_credito(start_date=fecha_inicio_anno, end_date=fecha_corte_dt)
    regenerar_resumenes_credito_debito()

    # Filtrar movimientos desde el inicio del año fiscal hasta la fecha de corte
    # Usar fecha_a_iso para compatibilidad con Python 3.14
    debitos = MovimientoUnificadoDebito.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio_anno),
        fecha__lte=fecha_a_iso(fecha_corte_dt)
    )
    creditos = MovimientoUnificadoCredito.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio_anno),
        fecha__lte=fecha_a_iso(fecha_corte_dt)
    )
    
    debitos_dict = {}
    for d in debitos:
        debitos_dict[d.cta_debito] = debitos_dict.get(d.cta_debito, 0) + float(d.monto_debito)
    creditos_dict = {}
    for c in creditos:
        creditos_dict[c.cta_credito] = creditos_dict.get(c.cta_credito, 0) + float(c.monto_credito)

    for fila in balance_rows:
        codigo = str(fila['codigo'])
        debito = debitos_dict.get(int(codigo), 0)
        credito = creditos_dict.get(int(codigo), 0)
        saldo_deudor = debito - credito if debito > credito else 0
        saldo_acreedor = credito - debito if credito > debito else 0
        if 1010100 <= int(codigo) <= 2040000:
            matriz_dict[f'A:{codigo}'] = saldo_deudor
            matriz_dict[f'P:{codigo}'] = saldo_acreedor
        elif 3010100 <= int(codigo) <= 3030300:
            matriz_dict[f'Pe:{codigo}'] = saldo_deudor
            matriz_dict[f'G:{codigo}'] = saldo_acreedor

    return render(request, "boran_app/resumen_financiero_segun_fecha.html", {
        'matriz_js': json.dumps(matriz_dict),
        'fecha_corte': fecha_corte,
        'panel_year': anno_fiscal,
    })

# vEXPORTAR A EXCEL

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from boran_app.resumen_financiero import RESUMEN_ACTIVO, RESUMEN_PASIVO, RESUMEN_RESULTADO
from boran_app.utils_financiero import eval_formula

def exportar_excel_resumen_financiero(request):
    # Supón que ya generaste matriz_dict en tu view principal según la fecha de corte
    matriz_dict = ...  # aquí debes obtenerla (¡esto es parte de tu lógica!)

    # 1. Calcula resultados en arrays:
    resultados_activo = {}
    for row in RESUMEN_ACTIVO:
        val = eval_formula(row["formula"], matriz_dict, resultados_activo)
        resultados_activo[str(row["linea"])] = val
        row["resultado"] = val

    resultados_resultado = {}
    for row in RESUMEN_RESULTADO:
        val = eval_formula(row["formula"], matriz_dict, resultados_resultado)
        resultados_resultado[str(row["linea"])] = val
        row["resultado"] = val

    resultados_pasivo = {}
    for row in RESUMEN_PASIVO:
        val = eval_formula(
            row["formula"], matriz_dict, resultados_pasivo,
            resultado_lookup=resultados_resultado,
            activo_lookup=resultados_activo
        )
        resultados_pasivo[str(row["linea"])] = val
        row["resultado"] = val

    # 2. Arma la tabla con tres bloques en la misma hoja:
    resumen = [
        ["Línea", "Cuenta Activo", "Monto Activo",
         "Línea", "Cuenta Pasivo", "Monto Pasivo",
         "Línea", "Cuenta Resultado", "Monto Resultado"]
    ]
    max_rows = max(len(RESUMEN_ACTIVO), len(RESUMEN_PASIVO), len(RESUMEN_RESULTADO))
    for i in range(max_rows):
        fila_activo = RESUMEN_ACTIVO[i] if i < len(RESUMEN_ACTIVO) else {"linea":"", "cuenta":"", "resultado":""}
        fila_pasivo = RESUMEN_PASIVO[i] if i < len(RESUMEN_PASIVO) else {"linea":"", "cuenta":"", "resultado":""}
        fila_resultado = RESUMEN_RESULTADO[i] if i < len(RESUMEN_RESULTADO) else {"linea":"", "cuenta":"", "resultado":""}
        resumen.append([
            fila_activo["linea"], fila_activo["cuenta"], fila_activo["resultado"],
            fila_pasivo["linea"], fila_pasivo["cuenta"], fila_pasivo["resultado"],
            fila_resultado["linea"], fila_resultado["cuenta"], fila_resultado["resultado"]
        ])

    # 3. Exporta a Excel PRO
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Financiero"

    # Cabecera pro
    header_font = Font(bold=True, size=13, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='B7B7B7'),
        right=Side(style='thin', color='B7B7B7'),
        top=Side(style='thin', color='B7B7B7'),
        bottom=Side(style='thin', color='B7B7B7')
    )
    ws.append(resumen[0])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
    ws.freeze_panes = 'A2'

    fill_even = PatternFill("solid", fgColor="E9F1FB")
    fill_odd = PatternFill("solid", fgColor="FFFFFF")
    for row_idx, row in enumerate(resumen[1:], start=2):
        ws.append(row)
        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
            cell.border = thin_border
            cell.fill = fill_even if row_idx % 2 == 0 else fill_odd

    for col_idx in range(1, len(resumen[0]) + 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in ws[get_column_letter(col_idx)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 45)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ResumenFinanciero.xlsx"'
    wb.save(response)
    return response

#----------------------------
#VENTAS CONSULTA EXPORTAR
#---------------------------


import pandas as pd
from django.http import HttpResponse
from .models import VentasConsulta

def exportar_ventasconsulta_excel(request):
    qs = VentasConsulta.objects.all().values()
    if not qs:
        return HttpResponse("No hay datos para exportar.")

    df = pd.DataFrame.from_records(qs)

    columnas = [
        'fecha', 'codigo_producto', 'comprador', 'cantidad', 'total_venta', 'cuenta_debito', 'debito',
        'cuenta_credito', 'cuenta_debito_eerr', 'debito_eerr', 'cuenta_credito_eerr', 'credito_eerr',
        'costo_directo_producto', 'comentario', 'costo_venta', 'categoria', 'producto', 
        'cuenta_debito_envio', 'credito_iva', 'venta_neta_iva', 'credito_envio', 'debito_envio'
    ]
    # Deja solo columnas válidas
    df = df[[col for col in columnas if col in df.columns]]

    # ---- Conversión de columnas numéricas ----
    columnas_int = [
        'cantidad'
        # agrega aquí otras columnas que DEBEN ser enteros (si aplica)
    ]
    columnas_float = [
        'total_venta', 'debito', 'debito_eerr', 'credito_eerr',
        'costo_directo_producto', 'costo_venta', 'credito_iva', 'venta_neta_iva',
        'credito_envio', 'debito_envio'
        # agrega aquí otras columnas DECIMALES si aplica
    ]
    for col in columnas_int:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')  # admite nulos
    for col in columnas_float:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').astype(float)

    # ---------------------------
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ventasconsulta.xlsx'
    df.to_excel(response, index=False)
    return response

# ——————————————————————————————————————————————————————————————
# RESUMEN MENSUAL
# ——————————————————————————————————————————————————————————————

def resumen_mensual(request):
    # Obtener fechas del año fiscal y regenerar tablas
    fecha_inicio, fecha_fin, anno_fiscal = regenerar_tablas_financieras(request)
   
    # Filtrar por año fiscal
    qs = (
        ResumenMensual.objects
        .filter(mes__year=anno_fiscal)
        .values('mes', 'ventas', 'costos', 'utilidad')
        .order_by('mes')
    )
    rows = list(qs)

    # ------- Utilidades de fecha -------
    def to_month_start(value):
        """Normaliza 'mes' a date(YYYY,MM,1) soportando date, datetime o str."""
        if value is None:
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return date(value.year, value.month, 1)
        if isinstance(value, datetime):
            d = value.date()
            return date(d.year, d.month, 1)
        if isinstance(value, str):
            for fmt in ("%Y-%m", "%Y-%m-%d"):
                try:
                    dt = datetime.strptime(value, fmt)
                    return date(dt.year, dt.month, 1)
                except ValueError:
                    continue
        raise ValueError(f"Formato de 'mes' no reconocido: {value!r}")

    def add_month(d: date) -> date:
        """Suma 1 mes a 'd' (primer día del mes)."""
        if d.month == 12:
            return date(d.year + 1, 1, 1)
        return date(d.year, d.month + 1, 1)

    # ------- Agregación por mes -------
    data_by_month = {}
    for r in rows:
        mes = to_month_start(r.get('mes'))
        if mes is None:
            continue
        ventas = Decimal(r.get('ventas') or 0)
        costos = Decimal(r.get('costos') or 0)
        utilidad = r.get('utilidad')
        utilidad = Decimal(utilidad) if utilidad is not None else (ventas - costos)

        if mes not in data_by_month:
            data_by_month[mes] = {'mes': mes, 'ventas': ventas, 'costos': costos, 'utilidad': utilidad}
        else:
            data_by_month[mes]['ventas']   += ventas
            data_by_month[mes]['costos']   += costos
            data_by_month[mes]['utilidad'] += utilidad

    # ------- Sin datos: contexto vacío seguro -------
    if not data_by_month:
        contexto = {
            'resumenes': [],
            'totales': {'ventas': Decimal('0'), 'costos': Decimal('0'), 'utilidad': Decimal('0')},
            'chart_data': {'labels': [], 'ventas': [], 'costos': [], 'utilidad': []},
        }
        return render(request, 'boran_app/resumen_mensual.html', contexto)

    # ------- Determinar rango: usar el año fiscal seleccionado -------
    months_sorted = sorted(data_by_month.keys())
    
    # Usar el año fiscal de la sesión
    start = date(anno_fiscal, 1, 1)
    end = date(anno_fiscal, 12, 31)
    
    # Si hay datos, ajustar el fin al último mes con datos (pero dentro del año fiscal)
    if months_sorted:
        ultimo_mes_datos = months_sorted[-1]
        if ultimo_mes_datos.year == anno_fiscal and ultimo_mes_datos < end:
            end = ultimo_mes_datos

    # Si por alguna razón 'end' quedó antes que 'start' (no debería), corrige:
    if end < start:
        end = start

    # ------- Reconstruir meses + utilidad acumulada -------
    resumenes_procesados = []
    cur = start
    running_utilidad = Decimal('0')
    while cur <= end:
        base = data_by_month.get(cur, {'mes': cur, 'ventas': Decimal('0'), 'costos': Decimal('0'), 'utilidad': Decimal('0')})
        # Copia para no mutar el diccionario original
        reg = {
            'mes': base['mes'],
            'ventas': Decimal(base['ventas']),
            'costos': Decimal(base['costos']),
            'utilidad': Decimal(base['utilidad']),
        }
        running_utilidad += reg['utilidad']
        reg['utilidad_acumulada'] = running_utilidad
        resumenes_procesados.append(reg)
        cur = add_month(cur)

    # ------- Totales del rango mostrado -------
    totales = {
        'ventas':   sum((r['ventas']   for r in resumenes_procesados), Decimal('0')),
        'costos':   sum((r['costos']   for r in resumenes_procesados), Decimal('0')),
        'utilidad': sum((r['utilidad'] for r in resumenes_procesados), Decimal('0')),
    }

    # ------- Datos para Chart.js -------
    chart_labels   = [f"{r['mes'].year:04d}-{r['mes'].month:02d}" for r in resumenes_procesados]
    chart_ventas   = [float(r['ventas'])   for r in resumenes_procesados]
    chart_costos   = [float(r['costos'])   for r in resumenes_procesados]
    chart_utilidad = [float(r['utilidad']) for r in resumenes_procesados]

    chart_data = {
        'labels': chart_labels,
        'ventas': chart_ventas,
        'costos': chart_costos,
        'utilidad': chart_utilidad,
    }

    return render(request, 'boran_app/resumen_mensual.html', {
        'resumenes': resumenes_procesados,
        'totales': totales,
        'chart_data': chart_data,
    })

#-----------------------------------------------
# ACTUALIZAR RESUMEN MENSUAL
#----------------------------------------------

CUENTAS_VENTAS = [3010101, 3010111]
CUENTA_COSTO = [3010200, 3010201, 3010202, 3010203,3010205,3010211,3010212,3010213,3010214,3010215,3010216, 3020200,3010300,3010400,3020500,3020600,3020700,3020800,3020900,3030100]

def actualizar_resumen_mensual(request):
    # Obtener fechas del año fiscal y regenerar tablas
    fecha_inicio, fecha_fin, anno_fiscal = regenerar_tablas_financieras(request)
    
    saldos_mensuales = defaultdict(lambda: defaultdict(lambda: {'debitos': 0, 'creditos': 0}))

    # Filtrar débitos por año fiscal
    # Usar fecha_a_iso para compatibilidad con Python 3.14
    debitos = MovimientoUnificadoDebito.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio),
        fecha__lte=fecha_a_iso(fecha_fin)
    ).annotate(
        mes=TruncMonth('fecha')
    ).values('mes', 'cta_debito').annotate(total=Sum('monto_debito'))

    for row in debitos:
        mes = row['mes']
        cuenta = str(row['cta_debito'])
        saldos_mensuales[mes][cuenta]['debitos'] += row['total'] or 0

    # Filtrar créditos por año fiscal
    # Usar fecha_a_iso para compatibilidad con Python 3.14
    creditos = MovimientoUnificadoCredito.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio),
        fecha__lte=fecha_a_iso(fecha_fin)
    ).annotate(
        mes=TruncMonth('fecha')
    ).values('mes', 'cta_credito').annotate(total=Sum('monto_credito'))

    for row in creditos:
        mes = row['mes']
        cuenta = str(row['cta_credito'])
        saldos_mensuales[mes][cuenta]['creditos'] += row['total'] or 0

    CUENTA_COSTO_STR = [str(c) for c in CUENTA_COSTO]
    CUENTAS_VENTAS_STR = [str(c) for c in CUENTAS_VENTAS]

    for mes, cuentas in saldos_mensuales.items():
        ventas = sum(
            cuentas.get(c, {}).get('creditos', 0) - cuentas.get(c, {}).get('debitos', 0)
            for c in CUENTAS_VENTAS_STR
        )
        costo_venta = sum(
            cuentas.get(c, {}).get('debitos', 0) - cuentas.get(c, {}).get('creditos', 0)
            for c in CUENTA_COSTO_STR
        )
        utilidad = ventas - costo_venta
        margen_bruto = utilidad

        ResumenMensual.objects.update_or_create(
            mes=mes,
            defaults={
                'ventas': ventas,
                'costos': costo_venta,
                'utilidad': utilidad,
                'margen_bruto': margen_bruto
            }
        )

    # --- CÁLCULO Y GUARDADO DE UTILIDAD ACUMULADA ---
    resumenes = ResumenMensual.objects.order_by('mes')
    acumulado = 0
    for r in resumenes:
        acumulado += r.utilidad or 0
        if r.utilidad_acumulada != acumulado:
            r.utilidad_acumulada = acumulado
            r.save(update_fields=['utilidad_acumulada'])

   #messages.success(request, "¡Resumen mensual actualizado con éxito!")
    return redirect('resumen_mensual')

# ——————————————————————————————————————————————————————————————
# RESULTADO DETALLADO POR MES
# ——————————————————————————————————————————————————————————————



CONCEPTOS_FINALES = [
    ("ventas", "Ventas"),
    ("costo_venta", "Costo de Venta"),
    ("margen_bruto", "Margen Bruto"),
    ("gastos_comercializacion", "Gastos Comercialización"),
    ("comision_plataformas", "Comision Plataforma de Pago"),
    ("gastos_marketing", "Gastos Publicidad y Marketing"),
    ("gastos_arriendos", "Gastos arriendos Comisiones Tiendas"),
    ("gastos_envios", "Gastos de Envíos Adicionales"),
    ("gastos_administracion", "Gastos de Administración"),
    ("resultado_operacional_bruto", "Resultado Operacional Bruto"),
    ("gastos_financieros", "Gastos Financieros"),
    ("depreciacion", "Depreciación del Ejercicio"),
    ("resultado_oper_neto", "Resultado Oper. Neto"),
    ("utilidad_no_operacional", "Utilidad (pérd.) No Operacional"),
    ("ajuste_monetario", "Ajuste Monetario"),
    ("impuesto_renta", "Impuesto a la Renta"),
    ("ajustes", "Ajustes"),
    ("utilidad_neta", "Utilidad Neta del Periodo"),
]

FILAS_CALCULADAS = [
    "margen_bruto",
    "resultado_operacional_bruto",
    "resultado_oper_neto",
    "utilidad_neta",
]

def tabla_resultados_mensual(request):

    # Obtener fechas del año fiscal y regenerar tablas
    fecha_inicio, fecha_fin, anno_fiscal = regenerar_tablas_financieras(request)

    año = anno_fiscal
    # Si es el año actual, usar hasta el mes actual. Si no, mostrar todos los 12 meses.
    if anno_fiscal == date.today().year:
        mes_actual = date.today().month
    else:
        mes_actual = 12

    detalles = ResultadoMensualDetalle.objects.filter(mes__year=año)
    datos = defaultdict(lambda: defaultdict(int))
    for d in detalles:
        datos[d.mes.month][d.concepto] = d.valor

    matriz = {c[0]: [None]*mes_actual for c in CONCEPTOS_FINALES}
    for mes in range(1, mes_actual + 1):
        base = datos[mes]
        margen_bruto = base.get("ventas", 0) - base.get("costo_venta", 0)
        gastos_suma = sum(base.get(k, 0) for k in [
            "gastos_comercializacion", "comision_plataformas","gastos_marketing", "gastos_arriendos",
            "gastos_envios", "gastos_administracion"
        ])
        resultado_operacional_bruto = margen_bruto - gastos_suma
        resultado_oper_neto = resultado_operacional_bruto - base.get("gastos_financieros", 0) - base.get("depreciacion", 0)
        utilidad_neta = (
            resultado_oper_neto
            - base.get("utilidad_no_operacional", 0)
            - base.get("ajuste_monetario", 0)
            - base.get("impuesto_renta", 0)
            - base.get("ajustes", 0)
        )
        matriz["ventas"][mes-1] = base.get("ventas")
        matriz["costo_venta"][mes-1] = base.get("costo_venta")
        matriz["margen_bruto"][mes-1] = margen_bruto
        matriz["gastos_comercializacion"][mes-1] = base.get("gastos_comercializacion")
        matriz["comision_plataformas"][mes-1] = base.get("comision_plataformas")
        matriz["gastos_marketing"][mes-1] = base.get("gastos_marketing")
        matriz["gastos_arriendos"][mes-1] = base.get("gastos_arriendos")
        matriz["gastos_envios"][mes-1] = base.get("gastos_envios")
        matriz["gastos_administracion"][mes-1] = base.get("gastos_administracion")
        matriz["resultado_operacional_bruto"][mes-1] = resultado_operacional_bruto
        matriz["gastos_financieros"][mes-1] = base.get("gastos_financieros")
        matriz["depreciacion"][mes-1] = base.get("depreciacion")
        matriz["resultado_oper_neto"][mes-1] = resultado_oper_neto
        matriz["utilidad_no_operacional"][mes-1] = base.get("utilidad_no_operacional")
        matriz["ajuste_monetario"][mes-1] = base.get("ajuste_monetario")
        matriz["impuesto_renta"][mes-1] = base.get("impuesto_renta")
        matriz["ajustes"][mes-1] = base.get("ajustes")
        matriz["utilidad_neta"][mes-1] = utilidad_neta

    MESES_ES = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    meses = MESES_ES[:mes_actual]

    filas = [
        (c[0], c[1], matriz[c[0]], sum(v for v in matriz[c[0]] if v is not None))
        for c in CONCEPTOS_FINALES
    ]

    return render(request, "boran_app/tabla_resultados_mensual.html", {
        "año": año,
        "meses": meses,
        "filas": filas,
        "filas_calculadas": FILAS_CALCULADAS,
    })



def actualizar_resultados_mensuales(request):
    calcular_resultados_mensuales()
    return redirect('tabla_resultados_mensual')  # o donde quieras redirigir

# ——————————————————————————————————————————————————————————————
# RESUMEN VENTAS POR TIENDAS
# ——————————————————————————————————————————————————————————————

def obtener_matriz_dict_con_request(request):
    """
    Versión de obtener_matriz_dict que usa el año fiscal de la sesión.
    """
    # Obtener fechas del año fiscal y regenerar tablas
    fecha_inicio, fecha_fin, anno_fiscal = regenerar_tablas_financieras(request)

    debitos_dict = {d.cuenta_debito: float(d.total_debito) for d in ResumenDebito.objects.all()}
    creditos_dict = {c.cuenta_credito: float(c.total_credito) for c in ResumenCredito.objects.all()}
    matriz_dict = {}
    for fila in balance_rows:
        codigo = str(fila['codigo'])
        debito = debitos_dict.get(int(codigo), 0)
        credito = creditos_dict.get(int(codigo), 0)
        saldo_deudor = debito - credito if debito > credito else 0
        saldo_acreedor = credito - debito if credito > debito else 0
        if 1010100 <= int(codigo) <= 2040000:
            matriz_dict[f'A:{codigo}'] = saldo_deudor
            matriz_dict[f'P:{codigo}'] = saldo_acreedor
        elif 3010100 <= int(codigo) <= 3010300:
            matriz_dict[f'Pe:{codigo}'] = saldo_deudor
            matriz_dict[f'G:{codigo}'] = saldo_acreedor
    return matriz_dict

TIENDAS = [
    {"key": "online", "nombre": "Online", "gasto_key": "Pe:3010211"},
    {"key": "casa_moda", "nombre": "Casa Moda", "gasto_key": "Pe:3010212"},
    {"key": "casa_aura", "nombre": "Casa Aura", "gasto_key": "Pe:3010213"},
    {"key": "pucon", "nombre": "Pucon", "gasto_key": "Pe:3010216"},
    {"key": "uber_eats", "nombre": "Uber Eats", "gasto_key": None},
    {"key": "venta_manual", "nombre": "Venta Manual", "gasto_key": None},
]

@staff_member_required
def resumen_ventas_tiendas_view(request):
    # Obtener fechas del año fiscal
    fecha_inicio, fecha_fin, anno_fiscal = obtener_fechas_anno_fiscal(request)
    
    matriz_dict = obtener_matriz_dict_con_request(request)
    # Usar fecha_a_iso para compatibilidad con Python 3.14
    ventas = VentasConsulta.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio),
        fecha__lte=fecha_a_iso(fecha_fin)
    )

    # Inicializa datos para cada tienda
    datos_por_tienda = {}
    for tienda in TIENDAS:
        datos_por_tienda[tienda["key"]] = {
            "nombre": tienda["nombre"],
            "ventas_netas": 0,
            "costo_producto": 0,
            "gastos_tienda": 0,
        }

    total_ventas = 0
    total_costo = 0
    total_gastos = 0
    total_resultado = 0

    # Clasifica ventas en tiendas correctas
    for v in ventas:
        tienda_key = None
        comprador = v.comprador

        if str(comprador).strip().lower() == "pucon":
            tienda_key = "pucon"
        elif comprador in ["Casa Moda", "PARQUE ARAUCO"]:
            tienda_key = "casa_moda"
        elif es_online(comprador):
            tienda_key = "online"
        elif str(comprador).strip() == "Casa Aura":
            tienda_key = "casa_aura"
        elif str(comprador).strip() == "Uber Eats":
            tienda_key = "uber_eats"
        elif str(comprador).strip().lower() == "venta manual":
            tienda_key = "venta_manual"
        else:
            continue

        datos_por_tienda[tienda_key]["ventas_netas"] += float(v.venta_neta_iva or 0)
        datos_por_tienda[tienda_key]["costo_producto"] += float(v.costo_directo_producto or 0)
        total_ventas += float(v.venta_neta_iva or 0)
        total_costo += float(v.costo_directo_producto or 0)

    # Agrega gastos y calcula resultados para cada tienda
    lista_final = []
    for tienda in TIENDAS:
        key = tienda["key"]
        gasto_key = tienda.get("gasto_key")
        gastos = matriz_dict.get(gasto_key, 0) if gasto_key else 0
        datos_por_tienda[key]["gastos_tienda"] = gastos

        ventas_netas = datos_por_tienda[key]["ventas_netas"]
        costo_producto = datos_por_tienda[key]["costo_producto"]
        gastos_tienda = datos_por_tienda[key]["gastos_tienda"]
        resultado_directo = ventas_netas - costo_producto - gastos_tienda
        porcentaje_ventas = ventas_netas / total_ventas * 100 if total_ventas else 0
        porcentaje_resultado = resultado_directo / ventas_netas * 100 if ventas_netas else 0
        margen_bruto = round((ventas_netas / costo_producto), 1) if costo_producto else 0

        lista_final.append({
            "nombre": tienda["nombre"],
            "ventas_netas": ventas_netas,
            "costo_producto": costo_producto,
            "gastos_tienda": gastos_tienda,
            "resultado_directo": resultado_directo,
            "porcentaje_ventas": porcentaje_ventas,
            "porcentaje_resultado": porcentaje_resultado,
            "margen_bruto": margen_bruto,
        })

        total_gastos += gastos
        total_resultado += resultado_directo

    total_margen_bruto = round((total_ventas / total_costo), 1) if total_costo else 0
    total_porcentaje_resultado = total_resultado / total_ventas * 100 if total_ventas else 0

    return render(request, "boran_app/resumen_ventas_tiendas.html", {
        "datos_por_tienda": lista_final,
        "total_ventas": total_ventas,
        "total_costo": total_costo,
        "total_gastos": total_gastos,
        "total_resultado": total_resultado,
        "total_margen_bruto": total_margen_bruto,
        "total_porcentaje_resultado": total_porcentaje_resultado,
    })


# ——————————————————————————————————————————————————————————————
# EXPORTAR A EXCEL RESUMEN TIENDAS
# ——————————————————————————————————————————————————————————————

import pandas as pd
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from .models import VentasConsulta, ResumenDebito, ResumenCredito
from .cod_cuentas_balance import balance_rows

def es_online(comprador):
    if comprador is None:
        return True
    c = str(comprador).strip().lower()
    return c in ["shopify", "", "nan", "none"]

# (La función obtener_matriz_dict fue reemplazada por obtener_matriz_dict_con_request arriba)
# Las constantes TIENDAS están definidas más arriba en el archivo

@staff_member_required
def exportar_resumen_ventas_tiendas_excel(request):
    from django.utils import timezone

    # Obtener fechas del año fiscal
    fecha_inicio, fecha_fin, anno_fiscal = obtener_fechas_anno_fiscal(request)
    
    matriz_dict = obtener_matriz_dict_con_request(request)
    # Usar fecha_a_iso para compatibilidad con Python 3.14
    ventas = VentasConsulta.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio),
        fecha__lte=fecha_a_iso(fecha_fin)
    )

    rows = []
    total_ventas = total_costo = total_gastos = total_resultado = 0

    # Prepara estructura vacía para cada tienda
    datos_por_tienda = {}
    for tienda in TIENDAS:
        datos_por_tienda[tienda["key"]] = {
            "nombre": tienda["nombre"],
            "ventas_netas": 0,
            "costo_producto": 0,
            "gastos_tienda": 0,
        }

    # Clasifica ventas según la lógica correcta
    for v in ventas:
        tienda_key = None
        comprador = v.comprador

        if str(comprador).strip().lower() == "pucon":
            tienda_key = "pucon"
        elif comprador in ["Casa Moda", "PARQUE ARAUCO"]:
            tienda_key = "casa_moda"
        elif es_online(comprador):
            tienda_key = "online"
        elif str(comprador).strip() == "Casa Aura":
            tienda_key = "casa_aura"
        elif str(comprador).strip() == "Uber Eats":
            tienda_key = "uber_eats"
        elif str(comprador).strip().lower() == "venta manual":
            tienda_key = "venta_manual"
        else:
            continue

        datos_por_tienda[tienda_key]["ventas_netas"] += float(v.venta_neta_iva or 0)
        datos_por_tienda[tienda_key]["costo_producto"] += float(v.costo_directo_producto or 0)
        total_ventas += float(v.venta_neta_iva or 0)
        total_costo += float(v.costo_directo_producto or 0)

    # Ahora agrega gastos y genera rows para Excel
    for tienda in TIENDAS:
        key = tienda["key"]
        gasto_key = tienda.get("gasto_key")
        gastos = matriz_dict.get(gasto_key, 0) if gasto_key else 0
        datos_por_tienda[key]["gastos_tienda"] = gastos

        ventas_netas = datos_por_tienda[key]["ventas_netas"]
        costo_producto = datos_por_tienda[key]["costo_producto"]

        resultado_directo = ventas_netas - costo_producto - gastos

        # Margen Bruto como decimal con un decimal
        margen_bruto = round((ventas_netas / costo_producto), 1) if costo_producto else 0

        rows.append({
            "Tienda": tienda["nombre"],
            "Ventas Netas": ventas_netas,
            "Costo Producto": costo_producto,
            "Gastos Tienda": gastos,
            "Resultado Directo": resultado_directo,
            "Margen Bruto": margen_bruto,
        })

        total_gastos += gastos
        total_resultado += resultado_directo

    margen_bruto_total = round((total_ventas / total_costo), 1) if total_costo else 0
    rows.append({
        "Tienda": "TOTAL",
        "Ventas Netas": total_ventas,
        "Costo Producto": total_costo,
        "Gastos Tienda": total_gastos,
        "Resultado Directo": total_resultado,
        "Margen Bruto": margen_bruto_total,
    })

    # Ordena las filas: Online primero, el resto igual que tu vista principal
    def orden_fila(row):
        return (row["Tienda"] != "Online", row["Tienda"])
    rows = sorted(rows[:-1], key=orden_fila) + [rows[-1]]  # Totales al final

    df = pd.DataFrame(rows)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ResumenVentasTiendas_{timezone.now().date()}.xlsx'
    df.to_excel(response, index=False)
    return response

def importar_datos(request):
    return render(request, "boran_app/importar_datos.html")


# ——————————————————————————————————————————————————————————————
# VISTAS ADICIONALES - Productos Rentables, Inventario Tiendas, etc.
# ——————————————————————————————————————————————————————————————

def productos_rentables(request):
    """
    Vista para mostrar los productos más rentables según el año fiscal seleccionado.
    """
    # Obtener fechas del año fiscal
    fecha_inicio, fecha_fin, anno_fiscal = obtener_fechas_anno_fiscal(request)
    
    # Obtener ventas del año fiscal
    # Usar fecha_a_iso para compatibilidad con Python 3.14
    ventas = VentasConsulta.objects.filter(
        fecha__gte=fecha_a_iso(fecha_inicio),
        fecha__lte=fecha_a_iso(fecha_fin)
    ).values('codigo_producto', 'categoria', 'producto').annotate(
        total_cantidad=Sum('cantidad'),
        total_venta=Sum('total_venta'),
        total_costo=Sum('costo_venta')
    ).order_by('-total_venta')[:20]
    
    # Calcular rentabilidad
    productos = []
    for v in ventas:
        venta_total = float(v['total_venta'] or 0)
        costo_total = float(v['total_costo'] or 0)
        margen = venta_total - costo_total
        margen_pct = (margen / venta_total * 100) if venta_total > 0 else 0
        productos.append({
            'codigo': v['codigo_producto'],
            'categoria': v['categoria'],
            'producto': v['producto'],
            'cantidad': v['total_cantidad'],
            'venta_total': intdot(venta_total),
            'costo_total': intdot(costo_total),
            'margen': intdot(margen),
            'margen_pct': round(margen_pct, 1),
        })
    
    return render(request, 'boran_app/productos_rentables.html', {
        'productos': productos,
        'panel_year': anno_fiscal,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    })


def inventario_tiendas(request):
    """
    Vista para mostrar el inventario por tienda y SKU según el año fiscal seleccionado.
    """
    # Obtener fechas del año fiscal
    fecha_inicio, fecha_fin, anno_fiscal = obtener_fechas_anno_fiscal(request)
    
    from .models import Catalogo, InventarioInicial, EntradaProductos, Envios, Ventas, AjusteInventario, EnviosATiendas
    from django.db import models
    
    # Definir tiendas
    tiendas = ['Online', 'Casa Moda', 'Casa Aura', 'Pucon', 'Uber Eats']
    
    inventario = []
    for obj in Catalogo.objects.all():
        ini = InventarioInicial.objects.filter(sku=obj.sku).first()
        inicial = ini.stock if ini else 0
        
        # Usar fecha_a_iso para compatibilidad con Python 3.14
        ingresos = EntradaProductos.objects.filter(
            sku__sku=obj.sku,
            fecha__gte=fecha_a_iso(fecha_inicio),
            fecha__lte=fecha_a_iso(fecha_fin)
        ).aggregate(total=models.Sum('cantidad_ingresada'))['total'] or 0
        
        ventas = Ventas.objects.filter(
            sku__sku=obj.sku,
            fecha__gte=fecha_a_iso(fecha_inicio),
            fecha__lte=fecha_a_iso(fecha_fin)
        ).aggregate(total=models.Sum('cantidad'))['total'] or 0
        
        ajustes = AjusteInventario.objects.filter(
            sku__sku=obj.sku,
            fecha__gte=fecha_a_iso(fecha_inicio),
            fecha__lte=fecha_a_iso(fecha_fin)
        ).aggregate(total=models.Sum('cantidad'))['total'] or 0
        
        # Stock disponible
        stock = inicial + ingresos - ventas - ajustes
        
        if stock != 0:  # Solo mostrar productos con stock
            inventario.append({
                'sku': obj.sku,
                'categoria': obj.categoria,
                'producto': obj.producto,
                'inicial': inicial,
                'ingresos': ingresos,
                'ventas': ventas,
                'ajustes': ajustes,
                'stock': stock,
            })
    
    return render(request, 'boran_app/inventario_tiendas.html', {
        'inventario': inventario,
        'tiendas': tiendas,
        'panel_year': anno_fiscal,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    })


def validar_plan_cuentas(request):
    """
    Vista para validar cuentas contables ingresadas.
    """
    from .cod_cuentas_balance import balance_rows
    
    anno_fiscal = get_panel_year(request)
    
    # Obtener todas las cuentas usadas en los movimientos
    cuentas_debito = set(MovimientoUnificadoDebito.objects.values_list('cta_debito', flat=True))
    cuentas_credito = set(MovimientoUnificadoCredito.objects.values_list('cta_credito', flat=True))
    cuentas_usadas = cuentas_debito.union(cuentas_credito)
    
    # Obtener cuentas del plan de cuentas
    cuentas_plan = {str(row['codigo']) for row in balance_rows}
    
    # Encontrar cuentas no válidas
    cuentas_invalidas = []
    for cuenta in cuentas_usadas:
        if cuenta and str(cuenta) not in cuentas_plan:
            cuentas_invalidas.append(cuenta)
    
    return render(request, 'boran_app/validar_plan_cuentas.html', {
        'cuentas_invalidas': sorted(cuentas_invalidas),
        'total_cuentas_usadas': len(cuentas_usadas),
        'total_invalidas': len(cuentas_invalidas),
        'panel_year': anno_fiscal,
    })


def movimientos_cuenta(request):
    """
    Vista para ver movimientos de una cuenta específica.
    """
    anno_fiscal = get_panel_year(request)
    fecha_inicio = date(anno_fiscal, 1, 1)
    fecha_fin = date(anno_fiscal, 12, 31) if anno_fiscal != date.today().year else date.today()
    
    cuenta = request.GET.get('cuenta', '')
    movimientos_debito = []
    movimientos_credito = []
    
    if cuenta:
        # Usar fecha_a_iso para compatibilidad con Python 3.14
        movimientos_debito = MovimientoUnificadoDebito.objects.filter(
            cta_debito=cuenta,
            fecha__gte=fecha_a_iso(fecha_inicio),
            fecha__lte=fecha_a_iso(fecha_fin)
        ).order_by('fecha')
        
        movimientos_credito = MovimientoUnificadoCredito.objects.filter(
            cta_credito=cuenta,
            fecha__gte=fecha_a_iso(fecha_inicio),
            fecha__lte=fecha_a_iso(fecha_fin)
        ).order_by('fecha')
    
    return render(request, 'boran_app/movimientos_cuenta.html', {
        'cuenta': cuenta,
        'movimientos_debito': movimientos_debito,
        'movimientos_credito': movimientos_credito,
        'panel_year': anno_fiscal,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    })


def movimientos_por_fecha(request):
    """
    Vista para ver todos los movimientos de una fecha específica.
    """
    anno_fiscal = get_panel_year(request)
    
    fecha_str = request.GET.get('fecha', '')
    fecha_dt = None
    movimientos_debito = []
    movimientos_credito = []
    
    if fecha_str:
        # Parsear fecha en formato DD-MM-AA o DD-MM-YYYY
        try:
            from datetime import datetime
            for fmt in ("%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    fecha_dt = datetime.strptime(fecha_str, fmt).date()
                    break
                except ValueError:
                    continue
            
            if fecha_dt:
                # Usar fecha_a_iso para compatibilidad con Python 3.14
                movimientos_debito = MovimientoUnificadoDebito.objects.filter(
                    fecha=fecha_a_iso(fecha_dt)
                ).order_by('cta_debito')
                
                movimientos_credito = MovimientoUnificadoCredito.objects.filter(
                    fecha=fecha_a_iso(fecha_dt)
                ).order_by('cta_credito')
        except Exception:
            pass
    
    return render(request, 'boran_app/movimientos_por_fecha.html', {
        'fecha': fecha_dt,
        'fecha_str': fecha_str,
        'movimientos_debito': movimientos_debito,
        'movimientos_credito': movimientos_credito,
        'panel_year': anno_fiscal,
    })


def movimientos_por_rango(request):
    """
    Vista para ver todos los movimientos en un rango de fechas.
    """
    anno_fiscal = get_panel_year(request)
    
    desde_str = request.GET.get('desde', '')
    hasta_str = request.GET.get('hasta', '')
    desde_dt = None
    hasta_dt = None
    movimientos_debito = []
    movimientos_credito = []
    
    def parse_fecha(fecha_str):
        from datetime import datetime
        for fmt in ("%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(fecha_str, fmt).date()
            except ValueError:
                continue
        return None
    
    if desde_str and hasta_str:
        desde_dt = parse_fecha(desde_str)
        hasta_dt = parse_fecha(hasta_str)
        
        if desde_dt and hasta_dt:
            # Usar fecha_a_iso para compatibilidad con Python 3.14
            movimientos_debito = MovimientoUnificadoDebito.objects.filter(
                fecha__gte=fecha_a_iso(desde_dt),
                fecha__lte=fecha_a_iso(hasta_dt)
            ).order_by('fecha', 'cta_debito')
            
            movimientos_credito = MovimientoUnificadoCredito.objects.filter(
                fecha__gte=fecha_a_iso(desde_dt),
                fecha__lte=fecha_a_iso(hasta_dt)
            ).order_by('fecha', 'cta_credito')
    
    # Calcular totales
    total_debito = sum(float(m.monto_debito or 0) for m in movimientos_debito)
    total_credito = sum(float(m.monto_credito or 0) for m in movimientos_credito)
    
    return render(request, 'boran_app/movimientos_por_rango.html', {
        'desde': desde_dt,
        'hasta': hasta_dt,
        'desde_str': desde_str,
        'hasta_str': hasta_str,
        'movimientos_debito': movimientos_debito,
        'movimientos_credito': movimientos_credito,
        'total_debito': intdot(total_debito),
        'total_credito': intdot(total_credito),
        'panel_year': anno_fiscal,
    })
