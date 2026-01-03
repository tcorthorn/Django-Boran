from django.shortcuts import render
from django.db.models import Q, Sum, F, FloatField, ExpressionWrapper
from django.core.paginator import Paginator
from .models import ProductoRentable
from boran_app.utils import (regenerar_ventas_consulta,poblar_movimientos_unificados_credito,poblar_movimientos_unificados_debito,
    regenerar_resumenes_credito_debito,)
from django.db.models import Value, IntegerField, Case, When
from django.db.models.functions import Coalesce, Lower, Trim
import pandas as pd
from datetime import date, datetime
from django.http import HttpResponse
from boran_app.models import Catalogo, EntradaProductos, Ventas 
from .models import BodegaTienda,EnviosATiendas,InventarioInicialTiendas,AjusteInventarioTienda
from django.db.models import Case, When
from decimal import Decimal
from boran_app.models import VentasConsulta, ResumenCredito,ResumenDebito
from pathlib import Path
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from boran_app.scripts.import_inventario_inicial_tiendas import run_import as importar_inventario_from_xlsx
from django.shortcuts import redirect
from django.contrib import messages
from consult_app.validar_plan_cuentas import validar_plan_cuentas
from boran_app.utils import regenerar_resumenes_credito_debito
from django.apps import apps   # ⬅️ ESTA LÍNEA FALTA
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


#_____________________________________
# PRODUCTOS MAS RENTABLES
#_____________________________________

def productos_rentables(request):
    # ---- Recalcula automáticamente cada vez que visitas la página ----
    ProductoRentable.objects.all().delete()

    # Procesos previos (puedes incluir solo los necesarios para el financiero)
    regenerar_ventas_consulta()
    poblar_movimientos_unificados_debito()
    poblar_movimientos_unificados_credito()
    regenerar_resumenes_credito_debito()

    #ventas = VentasConsulta.objects.annotate(
        #venta_neta_total_fila=ExpressionWrapper(F('cantidad') * F('venta_neta_iva'), output_field=FloatField()),
        #costo_total_fila=ExpressionWrapper(F('cantidad') * F('costo_directo_producto'), output_field=FloatField()),
    #)

    #DEVOLUCIONES

    from django.db.models import F, FloatField, Value, ExpressionWrapper, Case, When

    ventas = (
        VentasConsulta.objects
        .annotate(
            venta_neta_total_fila=Case(
                When(
                    cantidad__lt=0,
                    then=ExpressionWrapper(
                        F('cantidad') * F('venta_neta_iva') * Value(-1),
                        output_field=FloatField()
                    )
                ),
                default=ExpressionWrapper(
                    F('cantidad') * F('venta_neta_iva'),
                    output_field=FloatField()
                ),
                output_field=FloatField(),
            ),
            costo_total_fila=Case(
                When(
                    cantidad__lt=0,
                    then=ExpressionWrapper(
                        F('cantidad') * F('costo_directo_producto') * Value(-1),
                        output_field=FloatField()
                    )
                ),
                default=ExpressionWrapper(
                    F('cantidad') * F('costo_directo_producto'),
                    output_field=FloatField()
                ),
                output_field=FloatField(),
            ),
        )
    )


    resumen = ventas.values('codigo_producto', 'categoria', 'producto').annotate(
        cantidad=Sum('cantidad'),
        venta_total=Sum('venta_neta_total_fila'),
        costo_total=Sum('costo_total_fila')
    )

    for v in resumen:
        venta_total = v['venta_total'] or 0
        costo_total = v['costo_total'] or 0
        utilidad = venta_total - costo_total
        # Margen bruto según tu requerimiento: venta_total / costo_total si costo_total > 0
        margen = (venta_total / costo_total) if costo_total > 0 else 0

        ProductoRentable.objects.create(
            codigo_producto=v['codigo_producto'],
            categoria=v['categoria'],
            producto=v['producto'],
            cantidad=v['cantidad'],
            venta_total=venta_total,
            costo_total=costo_total,
            utilidad_bruta_total=utilidad,
            margen_bruto=margen
        )

    # --- Búsqueda ---
    q = request.GET.get('q', '').strip()
    productos = ProductoRentable.objects.all()
    if q:
        productos = productos.filter(
            Q(codigo_producto__icontains=q) |
            Q(categoria__icontains=q) |
            Q(producto__icontains=q)
        )

    # --- Ordenamiento ---
    sort = request.GET.get('sort', 'venta_total')
    direction = request.GET.get('dir', 'desc')
    allowed_sorts = [
        'codigo_producto', 'categoria', 'producto', 'cantidad',
        'venta_total', 'costo_total', 'utilidad_bruta_total', 'margen_bruto'
    ]
    if sort not in allowed_sorts:
        sort = 'venta_total'
    order_by = sort if direction == 'asc' else f'-{sort}'
    productos = productos.order_by(order_by)

    # --- Paginación ---
    paginator = Paginator(productos, 25)  # 25 productos por página
    page_number = request.GET.get('page')
    productos_page = paginator.get_page(page_number)

    # --- Totales de la página actual ---
    total_cantidad = sum([p.cantidad or 0 for p in productos_page])
    total_venta = sum([p.venta_total or 0 for p in productos_page])
    total_costo = sum([p.costo_total or 0 for p in productos_page])
    total_utilidad = sum([p.utilidad_bruta_total or 0 for p in productos_page])

    # --- Totales globales del filtro ---
    global_total = productos.aggregate(
        global_total_cantidad=Sum('cantidad'),
        global_total_venta=Sum('venta_total'),
        global_total_costo=Sum('costo_total'),
        global_total_utilidad=Sum('utilidad_bruta_total'),
    )

    return render(request, 'consult_app/productos_rentables.html', {
        'productos_page': productos_page,
        'q': q,
        'sort': sort,
        'direction': direction,
        'total_cantidad': total_cantidad,
        'total_venta': total_venta,
        'total_costo': total_costo,
        'total_utilidad': total_utilidad,
        'global_total_cantidad': global_total['global_total_cantidad'] or 0,
        'global_total_venta': global_total['global_total_venta'] or 0,
        'global_total_costo': global_total['global_total_costo'] or 0,
        'global_total_utilidad': global_total['global_total_utilidad'] or 0,
    })

def exportar_productos_excel(request):
    productos = ProductoRentable.objects.all().values()
    df = pd.DataFrame(list(productos))
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos_rentables.xlsx'
    df.to_excel(response, index=False)
    return response

#_____________________________________
# NUEVA CONSULTA INVENTARIO
#_____________________________________

# IDs fijos provistos por ti
OFICINA_ID = 1
BODEGA_ID = 5  # Bodega es una tienda más (vende online)

def _parse_fecha_corte(request):

    regenerar_ventas_consulta()
    poblar_movimientos_unificados_debito()
    poblar_movimientos_unificados_credito()
    regenerar_resumenes_credito_debito()

    qs = request.GET.get("fecha_corte")
    if not qs:
        return date.today()
    try:
        return datetime.strptime(qs, "%Y-%m-%d").date()
    except ValueError:
        return date.today()

def _calcular_filas(fecha_corte):
    # Tiendas
    tiendas = list(BodegaTienda.objects.all().values("id", "nombre"))
    tienda_by_id = {t["id"]: t["nombre"] for t in tiendas}

    # Inventario inicial
    ini_qs = (InventarioInicialTiendas.objects
              .filter(fecha__lte=fecha_corte)
              .values("tienda_id", "sku__sku")
              .annotate(stock_inicial=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField()))))
    inicial = {(r["tienda_id"], r["sku__sku"]): r["stock_inicial"] for r in ini_qs}

    # Envíos desde Oficina a tiendas (recibido por tienda)
    envios_qs = (EnviosATiendas.objects
                 .filter(fecha__lte=fecha_corte)
                 .values("tienda_bodega_id", "sku__sku")
                 .annotate(recibido=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField()))))
    recibido_por_tienda = {(r["tienda_bodega_id"], r["sku__sku"]): r["recibido"] for r in envios_qs}

    # Total enviado por Oficina (para fila de Oficina)
    enviados_oficina_qs = (EnviosATiendas.objects
                           .filter(fecha__lte=fecha_corte)
                           .values("sku__sku")
                           .annotate(enviado=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField()))))
    enviado_oficina_por_sku = {r["sku__sku"]: r["enviado"] for r in enviados_oficina_qs}

    # Entradas de producto en Oficina
    entradas_oficina_qs = (EntradaProductos.objects
                           .filter(fecha__lte=fecha_corte)
                           .values("sku__sku")
                           .annotate(entradas=Coalesce(Sum("cantidad_ingresada"), Value(0, output_field=IntegerField()))))
    recibido_oficina_por_sku = {r["sku__sku"]: r["entradas"] for r in entradas_oficina_qs}

    
    nombres_ref = ["Oficina", "Casa Moda", "Casa Aura", "Pucon", "Bodega", "Otro"]
    tienda_id_map = {
        n: BodegaTienda.objects.filter(nombre__iexact=n).values_list("id", flat=True).first()
        for n in nombres_ref
    }
    OFICINA_ID   = tienda_id_map.get("Oficina")   or 1
    CASA_MODA_ID = tienda_id_map.get("Casa Moda") or 2
    CASA_AURA_ID = tienda_id_map.get("Casa Aura") or 3
    PUCON_ID     = tienda_id_map.get("Pucon")     or 4
    BODEGA_ID    = tienda_id_map.get("Bodega")    or 5   # ⚠️ Si tu Bodega es 6, al no encontrarla por nombre quedará 5
    OTRO_ID      = tienda_id_map.get("Otro")      or 6

   
    ventas_qs = (
        Ventas.objects.filter(fecha__lte=fecha_corte)
        .annotate(comp_norm=Lower(Trim("comprador")))
        .annotate(
            tienda_id=Case(
                When(comprador__isnull=True, then=Value(BODEGA_ID)),                  # Bodega (nulo)
                When(comp_norm__in=["", "nan", "shopify"], then=Value(BODEGA_ID)),    # Bodega
                When(
                Q(comp_norm__in=["casa moda"]) |
                Q(comp_norm__contains="parque arauco") |
                Q(comp_norm__contains="parquearauco"),
                then=Value(CASA_MODA_ID),
            ),
                When(comp_norm="casa aura", then=Value(CASA_AURA_ID)),
                When(comp_norm="pucon", then=Value(PUCON_ID)),
                default=Value(OTRO_ID),  # Resto
                output_field=IntegerField(),
            )
        )
        .values("tienda_id", "sku__sku")
        .annotate(ventas=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField())))
    )

    ventas_por_tienda = {(r["tienda_id"], r["sku__sku"]): r["ventas"] for r in ventas_qs}

    # Ajustes por tienda
    ajustes_qs = (AjusteInventarioTienda.objects
                  .filter(fecha__lte=fecha_corte)
                  .values("tienda_id", "sku__sku")
                  .annotate(ajustes=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField()))))
    ajustes_por_tienda = {(r["tienda_id"], r["sku__sku"]): r["ajustes"] for r in ajustes_qs}

    # Claves universo
    claves = set()
    claves.update(inicial.keys())
    claves.update(recibido_por_tienda.keys())
    claves.update(ventas_por_tienda.keys())
    claves.update(ajustes_por_tienda.keys())
    for sku in set(enviado_oficina_por_sku.keys()).union(recibido_oficina_por_sku.keys()):
        claves.add((OFICINA_ID, sku))

    # Info de producto
    todos_skus = sorted({sku for (_tid, sku) in claves})
    info_prod = {
    c["sku"]: c for c in Catalogo.objects.filter(sku__in=todos_skus).values(
        "sku", "producto", "categoria", "costo_directo_producto"
    )
    }

    # Filas detalle
    filas = []
    for (tid, sku) in sorted(claves, key=lambda x: (x[0], x[1])):
        es_oficina = (tid == OFICINA_ID)
        stock_inicial = inicial.get((tid, sku), 0)
        ventas = ventas_por_tienda.get((tid, sku), 0)
        ajustes = ajustes_por_tienda.get((tid, sku), 0)

        if es_oficina:
            recibido = recibido_oficina_por_sku.get(sku, 0)  # entradas
            enviado = enviado_oficina_por_sku.get(sku, 0)    # envíos a tiendas
            stock_actual = stock_inicial + recibido - enviado + ajustes
        else:
            recibido = recibido_por_tienda.get((tid, sku), 0)  # envíos desde oficina
            enviado = 0
            stock_actual = stock_inicial + recibido - ventas + ajustes

        prod = info_prod.get(sku, {})
        costo_unitario = Decimal(prod.get("costo_directo_producto") or 0)
        valor_inventario = Decimal(stock_actual) * costo_unitario


        filas.append({
            "tienda_id": tid,
            "tienda": tienda_by_id.get(tid, f"ID {tid}"),
            "sku": sku,
            "producto": prod.get("producto") or "",
            "categoria": prod.get("categoria") or "",
            "stock_inicial": stock_inicial,
            "enviado": enviado,
            "recibido": recibido,
            "ventas": ventas,
            "ajustes": ajustes,
            "stock_actual": stock_actual,
            "costo_unitario": costo_unitario,          # <-- nuevo
            "valor_inventario": valor_inventario,      # <-- nuevo
        })

    return filas, tiendas

def informe_inventario_tiendas(request):
    fecha_corte = _parse_fecha_corte(request)
    filas, tiendas = _calcular_filas(fecha_corte)

    # --- Entradas de productos por SKU (para el consolidado) ---
    entradas_qs = (
        EntradaProductos.objects
        .filter(fecha__lte=fecha_corte)
        .values("sku__sku")
        .annotate(entradas=Coalesce(Sum("cantidad_ingresada"), Value(0, output_field=IntegerField())))
    )
    entradas_por_sku = {r["sku__sku"]: r["entradas"] for r in entradas_qs}

    # --- Consolidado por SKU (sin columna 'enviado' y 'recibido' = entradas de productos) ---
    consolidado_map = {}
    for f in filas:
        sku = f["sku"]
        agg = consolidado_map.get(sku)
        if not agg:
            agg = consolidado_map[sku] = {
                "sku": sku,
                "producto": f["producto"],
                "categoria": f["categoria"],
                "stock_inicial": 0,
                # "enviado": 0,           # <- ya no se usa en el consolidado
                "recibido": 0,            # se setea más abajo con entradas_por_sku
                "ventas": 0,
                "ajustes": 0,
                "stock_actual": 0,
                "costo_unitario": f.get("costo_unitario", Decimal(0)),
                "valor_inventario": Decimal(0),
            }
        agg["stock_inicial"] += f["stock_inicial"]
        # agg["enviado"]       += f["enviado"]   # <- NO acumular
        # agg["recibido"]      += f["recibido"]  # <- NO acumular, se fija con entradas
        agg["ventas"]        += f["ventas"]
        agg["ajustes"]       += f["ajustes"]
        agg["stock_actual"]  += f["stock_actual"]

    # Fijar 'recibido' = Entradas de Productos y calcular valor inventario
    for agg in consolidado_map.values():
        agg["recibido"] = entradas_por_sku.get(agg["sku"], 0)
        agg["valor_inventario"] = Decimal(agg["stock_actual"]) * agg["costo_unitario"]

    consolidado = sorted(consolidado_map.values(), key=lambda x: x["sku"])

    # Totales generales (sin cambios)
    total_detalle_stock = sum(int(f["stock_actual"]) for f in filas)
    total_detalle_valor = sum(Decimal(f["valor_inventario"]) for f in filas)
    total_consol_stock = sum(int(r["stock_actual"]) for r in consolidado)
    total_consol_valor = sum(Decimal(r["valor_inventario"]) for r in consolidado)

    context = {
        "fecha_corte": fecha_corte,
        "OFICINA_ID": OFICINA_ID,
        "BODEGA_ID": BODEGA_ID,
        "tiendas": tiendas,
        "filas": filas,
        "consolidado": consolidado,
        "sku_list": sorted({f["sku"] for f in filas}),
        "categoria_list": sorted({f["categoria"] for f in filas if f["categoria"]}),
        "total_detalle_stock": total_detalle_stock,
        "total_detalle_valor": total_detalle_valor,
        "total_consol_stock": total_consol_stock,
        "total_consol_valor": total_consol_valor,
        # --- Nuevos flags para la plantilla del consolidado ---
        "ocultar_enviado_consol": True,
        "titulo_consolidado": "Consolidado por SKU (Recibido = Entradas de Productos)",
        "etiqueta_recibido_consol": "Entradas de Productos",
    }
    return render(request, "consult_app/inventario_tiendas.html", context)

# ----------------------------
# Exportación a Excel (xlsx)
# ----------------------------
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
def exportar_inventario_tiendas_excel(request):
    

    fecha_corte = _parse_fecha_corte(request)
    filas, _ = _calcular_filas(fecha_corte)

    # --- Consolidado por SKU (Recibido = Entradas de Productos; sin 'Enviado') ---
    entradas_qs = (
        EntradaProductos.objects
        .filter(fecha__lte=fecha_corte)
        .values("sku__sku")
        .annotate(entradas=Coalesce(Sum("cantidad_ingresada"), Value(0, output_field=IntegerField())))
    )
    entradas_por_sku = {r["sku__sku"]: r["entradas"] for r in entradas_qs}

    consolidado_map = {}
    for f in filas:
        sku = f["sku"]
        if sku not in consolidado_map:
            consolidado_map[sku] = {
                "sku": sku,
                "producto": f["producto"],
                "categoria": f["categoria"],
                "stock_inicial": 0,
                "recibido": 0,  # será asignado con entradas
                "ventas": 0,
                "ajustes": 0,
                "stock_actual": 0,
                "costo_unitario": f.get("costo_unitario", Decimal(0)),
                "valor_inventario": Decimal(0),
            }
        agg = consolidado_map[sku]
        agg["stock_inicial"] += f["stock_inicial"]
        # agg["enviado"]       += f["enviado"]   # <- NO usar
        # agg["recibido"]      += f["recibido"]  # <- NO usar
        agg["ventas"]        += f["ventas"]
        agg["ajustes"]       += f["ajustes"]
        agg["stock_actual"]  += f["stock_actual"]

    for agg in consolidado_map.values():
        agg["recibido"] = entradas_por_sku.get(agg["sku"], 0)
        agg["valor_inventario"] = Decimal(agg["stock_actual"]) * agg["costo_unitario"]

    consolidado = sorted(consolidado_map.values(), key=lambda x: x["sku"])


    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Detalle por tienda"
    ws2 = wb.create_sheet("Consolidado por SKU")

    # Hoja 1
    headers1 = ["Tienda", "SKU", "Producto", "Categoría", "Stock Inicial",
                "Enviado (Oficina)", "Recibido", "Ventas", "Ajustes", "Stock Actual"]
    ws1.append(headers1)
    for f in filas:
        ws1.append([
            f["tienda"], f["sku"], f["producto"], f["categoria"],
            f["stock_inicial"], f["enviado"], f["recibido"],
            f["ventas"], f["ajustes"], f["stock_actual"]
        ])

   # Hoja 2 (Consolidado) — sin 'Enviado' y renombrando 'Recibido'
    headers2 = ["SKU", "Producto", "Categoría", "Stock Inicial",
                "Entradas de Productos", "Ventas", "Ajustes", "Stock Actual"]
    ws2.append(headers2)
    for r in consolidado:
        ws2.append([
            r["sku"], r["producto"], r["categoria"],
            r["stock_inicial"], r["recibido"],  # <- entradas
            r["ventas"], r["ajustes"], r["stock_actual"]
        ])

    # Autofit básico
    for ws in (ws1, ws2):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

    file_name = f"inventario_tiendas_{fecha_corte.isoformat()}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    wb.save(response)
    return response


# ----------------------------------------
# IMPORTACION iNVENTARIO INICIAL TIENDAS
# ---------------------------------------

DEFAULT_XLSX_PATH = r"C:\Users\tcort\OneDrive\BORANGORA\Django-Boran\Otros\Inventario Inicial Tiendas.xlsx"
DEFAULT_SHEET = None  # o "Hoja1"

@login_required
@staff_member_required
def importar_inventario_inicial_tiendas(request):
    if request.method != "POST":
        return redirect("home")  # ajusta el nombre de tu URL del Home

    try:
        creados, actualizados, omitidos = importar_inventario_from_xlsx(
            file_path=DEFAULT_XLSX_PATH,
            sheet_name=DEFAULT_SHEET,
            create_missing_tiendas=False,  # cambia a True si quieres crear tiendas faltantes
            strict_sku=False,              # True para abortar si viene un SKU no existente
            date_format="%Y-%m-%d",
        )
        messages.success(
            request,
            f"Inventario Inicial importado. Creados: {creados} | Actualizados: {actualizados} | Omitidos: {omitidos}"
        )
    except Exception as e:
        messages.error(request, f"Error al importar Inventario Inicial: {e}")

    return redirect("home")



    # VALIDAR CUENTAS CONTABLES
    #----------------------------

def validar_plan_cuentas_view(request):

    regenerar_ventas_consulta()
    poblar_movimientos_unificados_debito()
    poblar_movimientos_unificados_credito()
    regenerar_resumenes_credito_debito()

    # (Opcional) recalcular resúmenes antes de validar.
# Si no quieres recalcular cada vez, comenta la línea siguiente.
    try:
        regenerar_resumenes_credito_debito()
    except Exception as e:
        messages.warning(request, f"No se pudo recalcular resúmenes: {e}")

    validar_plan_cuentas(request)
    # Vuelve al Home (ajusta el nombre de la URL según tu proyecto)
    return redirect('home')  # o 'home'

from collections import defaultdict
from itertools import zip_longest
from io import BytesIO
from datetime import datetime, date
from decimal import Decimal
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
import boran_app.models as boran_models  # ✅ módulo completo, fuera de cualquier función


def movimientos_cuenta_view(request):
    cuenta_raw = (request.GET.get("cuenta") or "").strip()
    hubo_busqueda = bool(cuenta_raw)
    error_msg = ""
    rows = []
    total_debitos = 0
    total_creditos = 0
    saldo_cuenta = 0

    cuenta = None
    if cuenta_raw:
        try:
            cuenta = int(cuenta_raw)
        except ValueError:
            error_msg = "El código de cuenta debe ser numérico."

    if cuenta and not error_msg:
        debitos_qs = (
            boran_models.MovimientoUnificadoDebito.objects
            .filter(cta_debito=cuenta)
            .values("fecha", "cta_debito", "monto_debito", "texto_coment", "tabla_origen")
            .order_by("fecha", "id")
        )
        creditos_qs = (
            boran_models.MovimientoUnificadoCredito.objects
            .filter(cta_credito=cuenta)
            .values("fecha", "cta_credito", "monto_credito", "texto_coment", "tabla_origen")
            .order_by("fecha", "id")
        )

        total_debitos = (
            boran_models.MovimientoUnificadoDebito.objects
            .filter(cta_debito=cuenta)
            .aggregate(total=Sum("monto_debito"))["total"] or 0
        )
        total_creditos = (
            boran_models.MovimientoUnificadoCredito.objects
            .filter(cta_credito=cuenta)
            .aggregate(total=Sum("monto_credito"))["total"] or 0
        )

        # ---- NUEVO: saldo (débitos - créditos) ----
        try:
            saldo_cuenta = (total_debitos or 0) - (total_creditos or 0)
        except TypeError:
            # por si viniera algo que no sea numérico, fuerza a Decimal/float
            from decimal import Decimal
            saldo_cuenta = Decimal(total_debitos or 0) - Decimal(total_creditos or 0)



        # --- Emparejar por fecha y alinear por índice (evita líneas vacías) ---
        deb_por_fecha = defaultdict(list)
        for d in debitos_qs:
            k = d["fecha"] or date.min
            deb_por_fecha[k].append({
                "fecha_debito": d["fecha"],
                "cta_debito": d["cta_debito"],
                "monto_debito": d["monto_debito"],
                "coment_debito": d["texto_coment"],
                "origen_debito": d["tabla_origen"],
            })

        cred_por_fecha = defaultdict(list)
        for c in creditos_qs:
            k = c["fecha"] or date.min
            cred_por_fecha[k].append({
                "fecha_credito": c["fecha"],
                "cta_credito": c["cta_credito"],
                "monto_credito": c["monto_credito"],
                "coment_credito": c["texto_coment"],
                "origen_credito": c["tabla_origen"],
            })

        rows = []
        for k_fecha in sorted(set(deb_por_fecha.keys()) | set(cred_por_fecha.keys())):
            ld = deb_por_fecha.get(k_fecha, [])
            lc = cred_por_fecha.get(k_fecha, [])
            for d, c in zip_longest(ld, lc, fillvalue=None):
                row = {
                    "fecha_debito": None, "cta_debito": None, "monto_debito": None,
                    "coment_debito": None, "origen_debito": None,
                    "fecha_credito": None, "cta_credito": None, "monto_credito": None,
                    "coment_credito": None, "origen_credito": None,
                }
                if d: row.update(d)
                if c: row.update(c)
                rows.append(row)
        # --- fin emparejamiento por fecha ---

    context = {
        "hubo_busqueda": hubo_busqueda,
        "error_msg": error_msg,
        "cuenta_valor": cuenta_raw,
        "rows": rows,
        "total_debitos": total_debitos,
        "total_creditos": total_creditos,
        "saldo_cuenta": saldo_cuenta,   # <-- NUEVO
    }
    return render(request, "consult_app/movimientos_cuenta.html", context)


def movimientos_cuenta_endpoint(request):
    """
    Si ?export=excel -> genera y devuelve XLSX.
    En otro caso, delega a movimientos_cuenta_view (sin tocarla).
    """
    if (request.GET.get("export") or "").lower() != "excel":
        return movimientos_cuenta_view(request)

    # --------- MISMAS consultas que usa la view ---------
    cuenta_raw = (request.GET.get("cuenta") or "").strip()
    if not cuenta_raw.isdigit():
        return HttpResponse("El parámetro 'cuenta' es requerido y debe ser numérico.", status=400)
    cuenta = int(cuenta_raw)

    debitos_qs = (
        boran_models.MovimientoUnificadoDebito.objects
        .filter(cta_debito=cuenta)
        .values("fecha", "cta_debito", "monto_debito", "texto_coment", "tabla_origen")
        .order_by("fecha", "id")
    )
    creditos_qs = (
        boran_models.MovimientoUnificadoCredito.objects
        .filter(cta_credito=cuenta)
        .values("fecha", "cta_credito", "monto_credito", "texto_coment", "tabla_origen")
        .order_by("fecha", "id")
    )

    total_debitos = (
        boran_models.MovimientoUnificadoDebito.objects
        .filter(cta_debito=cuenta)
        .aggregate(total=Sum("monto_debito"))["total"] or 0
    )
    total_creditos = (
        boran_models.MovimientoUnificadoCredito.objects
        .filter(cta_credito=cuenta)
        .aggregate(total=Sum("monto_credito"))["total"] or 0
    )

    # --- Emparejar por fecha para el Excel (mismo orden que la tabla) ---
    deb_por_fecha = defaultdict(list)
    for d in debitos_qs:
        k = d["fecha"] or date.min
        deb_por_fecha[k].append({
            "fecha_debito": d["fecha"],
            "cta_debito": d["cta_debito"],
            "monto_debito": d["monto_debito"],
            "coment_debito": d["texto_coment"],
            "origen_debito": d["tabla_origen"],
        })

    cred_por_fecha = defaultdict(list)
    for c in creditos_qs:
        k = c["fecha"] or date.min
        cred_por_fecha[k].append({
            "fecha_credito": c["fecha"],
            "cta_credito": c["cta_credito"],
            "monto_credito": c["monto_credito"],
            "coment_credito": c["texto_coment"],
            "origen_credito": c["tabla_origen"],
        })

    rows = []
    for k_fecha in sorted(set(deb_por_fecha.keys()) | set(cred_por_fecha.keys())):
        ld = deb_por_fecha.get(k_fecha, [])
        lc = cred_por_fecha.get(k_fecha, [])
        for d, c in zip_longest(ld, lc, fillvalue=None):
            row = {
                "fecha_debito": None, "cta_debito": None, "monto_debito": None,
                "coment_debito": None, "origen_debito": None,
                "fecha_credito": None, "cta_credito": None, "monto_credito": None,
                "coment_credito": None, "origen_credito": None,
            }
            if d: row.update(d)
            if c: row.update(c)
            rows.append(row)
    # --- fin emparejamiento por fecha ---

    return export_movimientos_excel(
        cuenta=cuenta,
        rows=rows,
        total_debitos=total_debitos,
        total_creditos=total_creditos,
    )


# ---------- Exportador a Excel (compatible con la estructura de rows) ----------
def export_movimientos_excel(*, cuenta, rows, total_debitos, total_creditos):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    def _cell_value(v):
        if v in (None, "—", ""):
            return None
        if isinstance(v, Decimal):
            return float(v)
        if isinstance(v, datetime):
            return datetime(v.year, v.month, v.day, v.hour, v.minute, v.second)
        if isinstance(v, date):
            return v
        return v

    wb = Workbook()
    ws = wb.active
    ws.title = f"Cuenta {cuenta}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    t = ws.cell(row=1, column=1, value=f"Movimientos por Cuenta {cuenta}")
    t.font = Font(bold=True, size=14)
    t.alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    headers = [
        "Fecha", "Cuenta Débito", "Monto Débito", "Comentario", "Tabla Origen",
        "Fecha", "Cuenta Crédito", "Monto Crédito", "Comentario", "Tabla Origen",
    ]
    header_fill = PatternFill("solid", fgColor="E4EDF4")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    start_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    number_fmt = "#,##0"
    r = start_row + 1
    for row in rows:
        values = [
            row["fecha_debito"], row["cta_debito"], row["monto_debito"],
            row["coment_debito"], row["origen_debito"],
            row["fecha_credito"], row["cta_credito"], row["monto_credito"],
            row["coment_credito"], row["origen_credito"],
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=_cell_value(val))
            cell.border = border
            if c in (3, 8):
                cell.number_format = number_fmt
                cell.alignment = right
            elif c in (1, 6):
                cell.alignment = center
        r += 1

    ws.cell(row=r, column=2, value="Σ Débitos").font = Font(bold=True)
    td = ws.cell(row=r, column=3, value=_cell_value(total_debitos))
    td.number_format = number_fmt; td.font = Font(bold=True); td.alignment = right

    ws.cell(row=r, column=7, value="Σ Créditos").font = Font(bold=True)
    tc = ws.cell(row=r, column=8, value=_cell_value(total_creditos))
    tc.number_format = number_fmt; tc.font = Font(bold=True); tc.alignment = right

    # Auto-ancho
    for col in range(1, 11):
        max_len = 0
        for row_cells in ws.iter_rows(min_row=4, max_row=r, min_col=col, max_col=col):
            txt = "" if row_cells[0].value is None else str(row_cells[0].value)
            max_len = max(max_len, len(txt))
        ws.column_dimensions[chr(64 + col)].width = min(max(10, max_len + 2), 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.getvalue()

    filename = f"movimientos_cuenta_{cuenta}.xlsx"
    resp = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
    resp["Content-Length"] = str(len(data))
    resp["Cache-Control"] = "no-store"
    resp["X-Content-Type-Options"] = "nosniff"
    return resp

#______________________________
# CUENTAS CONTABLES POR FECHA
#----------------------------

import re
from datetime import date

def parse_fecha_es(fecha_str: str):

    regenerar_ventas_consulta()
    poblar_movimientos_unificados_debito()
    poblar_movimientos_unificados_credito()
    regenerar_resumenes_credito_debito()

    """
    Acepta:
      - 'DD-MM-AA' / 'DD-MM-AAAA'
      - 'DD/MM/AA' / 'DD/MM/AAAA'
      - 'DD.MM.AA' / 'DD.MM.AAAA'
      - guiones Unicode (– — − …)
      - sólo dígitos 'DDMMAA' o 'DDMMAAAA'
    Si es AA, se interpreta como 2000+AA.
    """
    if not fecha_str:
        return None
    s = str(fecha_str).strip()
    s = re.sub(r"[\u2010-\u2015\u2212\uFE58\uFE63\uFF0D]", "-", s)  # normaliza guiones raros
    s = s.replace("/", "-").replace(".", "-")
    s = re.sub(r"\s+", "", s)

    if re.fullmatch(r"\d{6}", s):
        s = f"{s[0:2]}-{s[2:4]}-{s[4:6]}"
    elif re.fullmatch(r"\d{8}", s):
        s = f"{s[0:2]}-{s[2:4]}-{s[4:8]}"

    m = re.fullmatch(r"(\d{1,2})-(\d{1,2})-(\d{2}|\d{4})", s)
    if not m:
        return None

    d, mth, yy = map(int, m.groups())
    if yy < 100:
        yy += 2000
    try:
        return date(yy, mth, d)
    except ValueError:
        return None




# Función movimientos_por_fecha

def movimientos_por_fecha_view(request):
    # Obtiene las clases de modelo SIEMPRE, sin depender de imports
    DebitoModel  = apps.get_model('boran_app', 'MovimientoUnificadoDebito')
    CreditoModel = apps.get_model('boran_app', 'MovimientoUnificadoCredito')
    # Si tu etiqueta de app fuera distinta (p.ej. 'boran_app'), cambia 'boran_app' por la etiqueta real en INSTALLED_APPS.

    fecha_raw = (request.GET.get("fecha") or "").strip()
    hubo_busqueda = bool(fecha_raw)
    error_msg = ""
    rows = []
    total_debitos = 0
    total_creditos = 0

    fecha_sel = parse_fecha_es(fecha_raw) if fecha_raw else None
    if fecha_raw and not fecha_sel:
        error_msg = "La fecha debe tener formato DD-MM-AA (o DD-MM-AAAA)."

    if fecha_sel and not error_msg:
        deb_qs = (
            DebitoModel.objects
            .filter(fecha=fecha_sel)
            .values("fecha", "cta_debito", "monto_debito", "texto_coment", "tabla_origen")
            .order_by("id")
        )
        cred_qs = (
            CreditoModel.objects
            .filter(fecha=fecha_sel)
            .values("fecha", "cta_credito", "monto_credito", "texto_coment", "tabla_origen")
            .order_by("id")
        )

        # añade estas dos líneas:
        debitos = list(deb_qs)
        creditos = list(cred_qs)

        total_debitos = (
            DebitoModel.objects
            .filter(fecha=fecha_sel)
            .aggregate(total=Sum("monto_debito"))["total"] or 0
        )
        total_creditos = (
            CreditoModel.objects
            .filter(fecha=fecha_sel)
            .aggregate(total=Sum("monto_credito"))["total"] or 0
        )

        for d in deb_qs:
            rows.append({
                "fecha_debito": d["fecha"],
                "cta_debito": d["cta_debito"],
                "monto_debito": d["monto_debito"],
                "coment_debito": d["texto_coment"],
                "origen_debito": d["tabla_origen"],
                "fecha_credito": None,
                "cta_credito": None,
                "monto_credito": None,
                "coment_credito": None,
                "origen_credito": None,
                "_tipo": "D",
            })

        for c in cred_qs:
            rows.append({
                "fecha_debito": None,
                "cta_debito": None,
                "monto_debito": None,
                "coment_debito": None,
                "origen_debito": None,
                "fecha_credito": c["fecha"],
                "cta_credito": c["cta_credito"],
                "monto_credito": c["monto_credito"],
                "coment_credito": c["texto_coment"],
                "origen_credito": c["tabla_origen"],
                "_tipo": "C",
            })

        rows.sort(key=lambda r: r["_tipo"])

    ctx = {
        "hubo_busqueda": hubo_busqueda,
        "error_msg": error_msg,
        "fecha_valor": fecha_raw,
        "rows": rows,
        "total_debitos": total_debitos,
        "total_creditos": total_creditos,
        "debitos": debitos,     
        "creditos": creditos,    
    }
    return render(request, "consult_app/movimientos_por_fecha.html", ctx)


#Exportar a Exel

def exportar_movimientos_fecha_excel(request):
    DebitoModel  = apps.get_model('boran_app', 'MovimientoUnificadoDebito')
    CreditoModel = apps.get_model('boran_app', 'MovimientoUnificadoCredito')

    fecha_raw = (request.GET.get("fecha") or "").strip()
    fecha_sel = parse_fecha_es(fecha_raw) if fecha_raw else None

    rows = []
    if fecha_sel:
        deb_qs = (
            DebitoModel.objects
            .filter(fecha=fecha_sel)
            .values("fecha", "cta_debito", "monto_debito", "texto_coment", "tabla_origen")
            .order_by("id")
        )
        cred_qs = (
            CreditoModel.objects
            .filter(fecha=fecha_sel)
            .values("fecha", "cta_credito", "monto_credito", "texto_coment", "tabla_origen")
            .order_by("id")
        )

        for d in deb_qs:
            rows.append([
                d["fecha"], d["cta_debito"], d["monto_debito"], d["texto_coment"], d["tabla_origen"],
                None, None, None, None, None
            ])
        for c in cred_qs:
            rows.append([
                None, None, None, None, None,
                c["fecha"], c["cta_credito"], c["monto_credito"], c["texto_coment"], c["tabla_origen"]
            ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos por fecha"

    headers = [
        "Fecha (Débito)", "Cuenta Débito", "Monto Débito", "Comentario (D)", "Tabla Origen (D)",
        "Fecha (Crédito)", "Cuenta Crédito", "Monto Crédito", "Comentario (C)", "Tabla Origen (C)",
    ]
    ws.append(headers)

    bold = Font(bold=True); center = Alignment(horizontal="center")
    for idx in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=idx)
        cell.font = bold
        cell.alignment = center
        try:
            col_letter = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[col_letter].width = 18
        except Exception:
            pass

    for r in rows:
        ws.append(r)

    safe_name = (fecha_raw or 'sin_fecha').replace('/', '-').replace('.', '-')
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="movimientos_por_fecha_{safe_name}.xlsx"'
    wb.save(resp)
    return resp

#______________________________
# CUENTAS CONTABLES POR RANGO
#----------------------------
from django.db.models import Sum, Count

def movimientos_por_rango_view(request):

    regenerar_ventas_consulta()
    poblar_movimientos_unificados_debito()
    poblar_movimientos_unificados_credito()
    regenerar_resumenes_credito_debito()

    DebitoModel  = apps.get_model('boran_app', 'MovimientoUnificadoDebito')
    CreditoModel = apps.get_model('boran_app', 'MovimientoUnificadoCredito')

    desde_raw = (request.GET.get("desde") or "").strip()
    hasta_raw = (request.GET.get("hasta") or "").strip()
    hubo_busqueda = bool(desde_raw or hasta_raw)
    error_msg = ""
    rows = []

    # Validación y parseo de fechas
    if hubo_busqueda:
        if not desde_raw or not hasta_raw:
            error_msg = "Debes ingresar ambas fechas (Desde y Hasta) en formato DD-MM-AA."
            desde = hasta = None
        else:
            desde = parse_fecha_es(desde_raw)
            hasta = parse_fecha_es(hasta_raw)
            if not desde or not hasta:
                error_msg = "Formato de fecha inválido. Usa DD-MM-AA (o DD-MM-AAAA)."
            else:
                # si el usuario las invierte, las ordenamos
                if desde > hasta:
                    desde, hasta = hasta, desde
    else:
        desde = hasta = None

    total_debitos = 0
    total_creditos = 0

    if desde and hasta and not error_msg:
        # Agregaciones por cuenta en el rango
        deb_agg = (DebitoModel.objects
                   .filter(fecha__gte=desde, fecha__lte=hasta)
                   .values("cta_debito")
                   .annotate(total_debito=Sum("monto_debito"),
                             n_debito=Count("id"))
                   .order_by())
        cred_agg = (CreditoModel.objects
                    .filter(fecha__gte=desde, fecha__lte=hasta)
                    .values("cta_credito")
                    .annotate(total_credito=Sum("monto_credito"),
                              n_credito=Count("id"))
                    .order_by())

        # Merge por cuenta
        index = {}
        for d in deb_agg:
            cta = d["cta_debito"]
            index[cta] = {
                "cuenta": cta,
                "total_debito": d["total_debito"] or 0,
                "n_debito": d["n_debito"] or 0,
                "total_credito": 0,
                "n_credito": 0,
            }
        for c in cred_agg:
            cta = c["cta_credito"]
            item = index.get(cta)
            if not item:
                item = {
                    "cuenta": cta,
                    "total_debito": 0,
                    "n_debito": 0,
                    "total_credito": 0,
                    "n_credito": 0,
                }
                index[cta] = item
            item["total_credito"] += c["total_credito"] or 0
            item["n_credito"] += c["n_credito"] or 0

        # Construye filas y totales globales
        for cta, item in index.items():
            saldo = (item["total_debito"] or 0) - (item["total_credito"] or 0)
            rows.append({
                "cuenta": cta,
                "total_debito": item["total_debito"] or 0,
                "n_debito": item["n_debito"] or 0,
                "total_credito": item["total_credito"] or 0,
                "n_credito": item["n_credito"] or 0,
                "saldo": saldo,
            })
            total_debitos += item["total_debito"] or 0
            total_creditos += item["total_credito"] or 0

        # Orden por cuenta (o si prefieres por |saldo|, cámbialo)
        rows.sort(key=lambda r: r["cuenta"])

    ctx = {
        "hubo_busqueda": hubo_busqueda,
        "error_msg": error_msg,
        "desde_valor": desde_raw,
        "hasta_valor": hasta_raw,
        "rows": rows,
        "total_debitos": total_debitos,
        "total_creditos": total_creditos,
        "saldo_neto": (total_debitos - total_creditos) if (desde and hasta and not error_msg) else 0,
    }
    return render(request, "consult_app/movimientos_por_rango.html", ctx)

# Exportar a Excel

def exportar_movimientos_rango_excel(request):
    DebitoModel  = apps.get_model('boran_app', 'MovimientoUnificadoDebito')
    CreditoModel = apps.get_model('boran_app', 'MovimientoUnificadoCredito')

    desde_raw = (request.GET.get("desde") or "").strip()
    hasta_raw = (request.GET.get("hasta") or "").strip()
    desde = parse_fecha_es(desde_raw) if desde_raw else None
    hasta = parse_fecha_es(hasta_raw) if hasta_raw else None
    if desde and hasta and desde > hasta:
        desde, hasta = hasta, desde

    rows = []
    total_debitos = 0
    total_creditos = 0

    if desde and hasta:
        deb_agg = (DebitoModel.objects
                   .filter(fecha__gte=desde, fecha__lte=hasta)
                   .values("cta_debito")
                   .annotate(total_debito=Sum("monto_debito"),
                             n_debito=Count("id"))
                   .order_by())
        cred_agg = (CreditoModel.objects
                    .filter(fecha__gte=desde, fecha__lte=hasta)
                    .values("cta_credito")
                    .annotate(total_credito=Sum("monto_credito"),
                              n_credito=Count("id"))
                    .order_by())

        index = {}
        for d in deb_agg:
            cta = d["cta_debito"]
            index[cta] = {
                "cuenta": cta,
                "total_debito": d["total_debito"] or 0,
                "n_debito": d["n_debito"] or 0,
                "total_credito": 0,
                "n_credito": 0,
            }
        for c in cred_agg:
            cta = c["cta_credito"]
            item = index.get(cta)
            if not item:
                item = {
                    "cuenta": cta,
                    "total_debito": 0,
                    "n_debito": 0,
                    "total_credito": 0,
                    "n_credito": 0,
                }
                index[cta] = item
            item["total_credito"] += c["total_credito"] or 0
            item["n_credito"] += c["n_credito"] or 0

        for cta, item in index.items():
            saldo = (item["total_debito"] or 0) - (item["total_credito"] or 0)
            rows.append([
                cta,
                item["total_debito"] or 0,
                item["n_debito"] or 0,
                item["total_credito"] or 0,
                item["n_credito"] or 0,
                saldo,
            ])
            total_debitos += item["total_debito"] or 0
            total_creditos += item["total_credito"] or 0

        # opcional: ordenar por cuenta asc
        rows.sort(key=lambda r: r[0])

    # Construir XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen por cuenta"

    headers = ["Cuenta", "Σ Débitos", "#Mov D", "Σ Créditos", "#Mov C", "Saldo (D - C)"]
    ws.append(headers)
    bold = Font(bold=True); center = Alignment(horizontal="center")
    for i in range(1, len(headers)+1):
        c = ws.cell(row=1, column=i); c.font = bold; c.alignment = center
        try:
            ws.column_dimensions[c.column_letter].width = 18
        except Exception:
            pass

    for r in rows:
        ws.append(r)

    # Totales al final
    if rows:
        ws.append([])
        ws.append(["TOTALES", total_debitos, None, total_creditos, None, total_debitos - total_creditos])

    safe_desde = (desde_raw or "").replace("/", "-").replace(".", "-")
    safe_hasta = (hasta_raw or "").replace("/", "-").replace(".", "-")
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="movimientos_por_rango_{safe_desde}_a_{safe_hasta}.xlsx"'
    wb.save(resp)
    return resp
